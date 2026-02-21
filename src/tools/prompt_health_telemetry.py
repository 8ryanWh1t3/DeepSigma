#!/usr/bin/env python3
"""Prompt health telemetry: track usage, detect drift, update PromptLibraryTable.

Collects telemetry from prompt usage logs, detects output drift by comparing
actual output structure against expected format, and updates DriftFlag and
PromptHealth in the Prompt OS workbook.

Usage:
    # Record a prompt usage event
    python -m tools.prompt_health_telemetry record \
        --prompt-id PRM-001 --success --rating 4

    # Analyze telemetry and detect drift
    python -m tools.prompt_health_telemetry analyze

    # Update workbook PromptLibraryTable from telemetry
    python -m tools.prompt_health_telemetry update-workbook \
        --workbook artifacts/excel/Coherence_Prompt_OS_v2.xlsx

    # Full report
    python -m tools.prompt_health_telemetry report
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

TELEMETRY_DIR = Path("data/prompt_telemetry")
TELEMETRY_LOG = TELEMETRY_DIR / "usage_log.csv"
DRIFT_LOG = TELEMETRY_DIR / "drift_log.csv"

TELEMETRY_HEADERS = [
    "timestamp",
    "prompt_id",
    "success",
    "rating",
    "model",
    "output_has_expected_sections",
    "notes",
]

DRIFT_HEADERS = [
    "timestamp",
    "prompt_id",
    "drift_type",
    "severity",
    "detail",
]

# Expected output sections per prompt (for drift detection)
EXPECTED_SECTIONS = {
    "PRM-001": [  # Executive Analysis
        "Executive Summary",
        "Recommended Action",
        "Facts",
        "Assumptions",
        "Failure Modes",
        "Next",
    ],
    "PRM-002": [  # Reality Assessment
        "actually happening",
        "story",
        "assumptions",
        "contradicts",
        "Drift Check",
        "grounded",
    ],
    "PRM-003": [  # Team Workbook Triage
        "TOP RISKS",
        "TOP ACTIONS",
        "SYSTEM OBSERVATIONS",
        "Seal",
    ],
    "PRM-004": [  # Assumption Audit
        "assumptions",
        "expir",
        "confidence",
    ],
}


def ensure_telemetry_dir():
    TELEMETRY_DIR.mkdir(parents=True, exist_ok=True)
    if not TELEMETRY_LOG.exists():
        with open(TELEMETRY_LOG, "w", newline="") as f:
            csv.writer(f).writerow(TELEMETRY_HEADERS)
    if not DRIFT_LOG.exists():
        with open(DRIFT_LOG, "w", newline="") as f:
            csv.writer(f).writerow(DRIFT_HEADERS)


def cmd_record(args) -> int:
    """Record a prompt usage event."""
    ensure_telemetry_dir()
    now = datetime.now(timezone.utc).isoformat()
    row = [
        now,
        args.prompt_id,
        "true" if args.success else "false",
        str(args.rating) if args.rating else "",
        args.model or "",
        "",
        args.notes or "",
    ]
    with open(TELEMETRY_LOG, "a", newline="") as f:
        csv.writer(f).writerow(row)
    print(f"Recorded: {args.prompt_id} success={args.success} rating={args.rating}")
    return 0


def cmd_check_drift(args) -> int:
    """Check an LLM output for structural drift against expected format."""
    ensure_telemetry_dir()
    prompt_id = args.prompt_id
    output_text = args.output_file.read_text() if args.output_file else args.output_text

    if not output_text:
        print("ERROR: Provide --output-text or --output-file")
        return 1

    expected = EXPECTED_SECTIONS.get(prompt_id, [])
    if not expected:
        print(f"WARN: No expected sections defined for {prompt_id}")
        return 0

    missing = []
    for section in expected:
        if not re.search(re.escape(section), output_text, re.IGNORECASE):
            missing.append(section)

    coverage = (len(expected) - len(missing)) / len(expected) * 100
    has_expected = len(missing) == 0

    # Determine drift severity
    if coverage >= 90:
        severity = "None"
    elif coverage >= 70:
        severity = "Minor"
    else:
        severity = "Major"

    # Record in telemetry
    now = datetime.now(timezone.utc).isoformat()
    with open(TELEMETRY_LOG, "a", newline="") as f:
        csv.writer(f).writerow([
            now, prompt_id, "true", "", "", str(has_expected).lower(),
            f"Coverage: {coverage:.0f}%"
        ])

    if missing:
        with open(DRIFT_LOG, "a", newline="") as f:
            csv.writer(f).writerow([
                now, prompt_id, "format_drift", severity,
                f"Missing sections: {', '.join(missing)} ({coverage:.0f}% coverage)"
            ])

    print(f"Prompt: {prompt_id}")
    print(f"Section coverage: {coverage:.0f}% ({len(expected) - len(missing)}/{len(expected)})")
    print(f"Drift severity: {severity}")
    if missing:
        print(f"Missing: {', '.join(missing)}")
    return 0 if severity == "None" else 1


def cmd_analyze(args) -> int:
    """Analyze telemetry log and compute health metrics per prompt."""
    ensure_telemetry_dir()
    if not TELEMETRY_LOG.exists():
        print("No telemetry data found.")
        return 0

    stats: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "successes": 0, "ratings": [], "drift_events": 0
    })

    with open(TELEMETRY_LOG, newline="") as f:
        for row in csv.DictReader(f):
            pid = row["prompt_id"]
            stats[pid]["total"] += 1
            if row["success"] == "true":
                stats[pid]["successes"] += 1
            if row["rating"]:
                try:
                    stats[pid]["ratings"].append(float(row["rating"]))
                except ValueError:
                    pass

    # Count drift events
    if DRIFT_LOG.exists():
        with open(DRIFT_LOG, newline="") as f:
            for row in csv.DictReader(f):
                pid = row["prompt_id"]
                if pid in stats:
                    stats[pid]["drift_events"] += 1

    print(f"{'PromptID':<12} {'Uses':>5} {'Success%':>9} {'AvgRating':>10} {'Drift':>6} {'Health':>7}")
    print("-" * 55)

    for pid, s in sorted(stats.items()):
        success_rate = (s["successes"] / s["total"] * 100) if s["total"] else 0
        avg_rating = sum(s["ratings"]) / len(s["ratings"]) if s["ratings"] else 0
        drift_flag = "Major" if s["drift_events"] >= 3 else ("Minor" if s["drift_events"] >= 1 else "None")
        drift_penalty = 20 if drift_flag == "Major" else (10 if drift_flag == "Minor" else 0)
        health = max(0, min(100, success_rate * 0.5 + avg_rating * 10 - drift_penalty))

        print(f"{pid:<12} {s['total']:>5} {success_rate:>8.0f}% {avg_rating:>10.1f} {drift_flag:>6} {health:>6.0f}")

    return 0


def cmd_update_workbook(args) -> int:
    """Update PromptLibraryTable in workbook from telemetry data."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. pip install openpyxl")
        return 1

    ensure_telemetry_dir()
    wb_path = args.workbook

    if not wb_path.exists():
        print(f"ERROR: Workbook not found: {wb_path}")
        return 1

    # Compute stats (same as analyze)
    stats: dict[str, dict] = defaultdict(lambda: {
        "total": 0, "successes": 0, "ratings": [], "drift_events": 0
    })

    if TELEMETRY_LOG.exists():
        with open(TELEMETRY_LOG, newline="") as f:
            for row in csv.DictReader(f):
                pid = row["prompt_id"]
                stats[pid]["total"] += 1
                if row["success"] == "true":
                    stats[pid]["successes"] += 1
                if row["rating"]:
                    try:
                        stats[pid]["ratings"].append(float(row["rating"]))
                    except ValueError:
                        pass

    if DRIFT_LOG.exists():
        with open(DRIFT_LOG, newline="") as f:
            for row in csv.DictReader(f):
                pid = row["prompt_id"]
                if pid in stats:
                    stats[pid]["drift_events"] += 1

    if not stats:
        print("No telemetry data to apply.")
        return 0

    wb = openpyxl.load_workbook(wb_path)
    ws = None
    for sheet in wb.worksheets:
        for tbl in sheet.tables.values():
            if tbl.displayName == "PromptLibraryTable":
                ws = sheet
                break
        if ws:
            break

    if not ws:
        print("ERROR: PromptLibraryTable not found in workbook")
        return 1

    # Find column indices
    headers = {cell.value: cell.column for cell in ws[1]}
    pid_col = headers.get("PromptID")
    usage_col = headers.get("UsageCount")
    success_col = headers.get("SuccessRate_pct")
    rating_col = headers.get("AvgRating_1to5")
    drift_col = headers.get("DriftFlag")
    last_used_col = headers.get("LastUsed")

    updated = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        pid = row[pid_col - 1].value if pid_col else None
        if pid and pid in stats:
            s = stats[pid]
            success_rate = int(s["successes"] / s["total"] * 100) if s["total"] else 0
            avg_rating = round(sum(s["ratings"]) / len(s["ratings"]), 1) if s["ratings"] else 0
            drift_flag = "Major" if s["drift_events"] >= 3 else ("Minor" if s["drift_events"] >= 1 else "None")

            if usage_col:
                row[usage_col - 1].value = (row[usage_col - 1].value or 0) + s["total"]
            if success_col and success_rate:
                row[success_col - 1].value = success_rate
            if rating_col and avg_rating:
                row[rating_col - 1].value = avg_rating
            if drift_col:
                row[drift_col - 1].value = drift_flag
            if last_used_col:
                row[last_used_col - 1].value = date.today().isoformat()
            updated += 1
            print(f"  Updated {pid}: usage +{s['total']}, success={success_rate}%, drift={drift_flag}")

    wb.save(wb_path)
    print(f"\nWorkbook updated: {updated} prompts modified")
    return 0


def cmd_report(args) -> int:
    """Print full telemetry report."""
    ensure_telemetry_dir()
    print("=== Usage Log ===")
    if TELEMETRY_LOG.exists():
        with open(TELEMETRY_LOG) as f:
            print(f.read())
    else:
        print("(empty)")

    print("\n=== Drift Log ===")
    if DRIFT_LOG.exists():
        with open(DRIFT_LOG) as f:
            print(f.read())
    else:
        print("(empty)")

    print("\n=== Health Analysis ===")
    return cmd_analyze(args)


def main() -> int:
    parser = argparse.ArgumentParser(description="Prompt health telemetry for Prompt OS v2")
    sub = parser.add_subparsers(dest="command")

    # record
    rec = sub.add_parser("record", help="Record a prompt usage event")
    rec.add_argument("--prompt-id", required=True, help="Prompt ID (e.g., PRM-001)")
    rec.add_argument("--success", action="store_true", help="Mark as successful")
    rec.add_argument("--no-success", dest="success", action="store_false")
    rec.add_argument("--rating", type=float, help="Quality rating (1-5)")
    rec.add_argument("--model", help="LLM model used")
    rec.add_argument("--notes", help="Additional notes")

    # check-drift
    drift = sub.add_parser("check-drift", help="Check LLM output for structural drift")
    drift.add_argument("--prompt-id", required=True, help="Prompt ID")
    drift.add_argument("--output-text", help="LLM output text to check")
    drift.add_argument("--output-file", type=Path, help="File containing LLM output")

    # analyze
    sub.add_parser("analyze", help="Analyze telemetry and compute health metrics")

    # update-workbook
    upd = sub.add_parser("update-workbook", help="Update PromptLibraryTable from telemetry")
    upd.add_argument("--workbook", type=Path,
                     default=Path("artifacts/excel/Coherence_Prompt_OS_v2.xlsx"))

    # report
    sub.add_parser("report", help="Print full telemetry report")

    args = parser.parse_args()

    if args.command == "record":
        return cmd_record(args)
    elif args.command == "check-drift":
        return cmd_check_drift(args)
    elif args.command == "analyze":
        return cmd_analyze(args)
    elif args.command == "update-workbook":
        return cmd_update_workbook(args)
    elif args.command == "report":
        return cmd_report(args)
    else:
        parser.print_help()
        return 0


if __name__ == "__main__":
    sys.exit(main())
