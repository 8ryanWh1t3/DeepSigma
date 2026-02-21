#!/usr/bin/env python3
"""Generate CareOps Pack Excel workbook for the Healthcare Lattice example.

Creates a multi-tab workbook with synthetic data aligned to the Meridian Health
Partners (fictional) multi-facility scenario. Each tab is a named Excel Table.

Requires: openpyxl (pip install openpyxl)

Usage:
    python ./examples/05-healthcare-lattice/tools/generate_healthcare_workbook.py
    python ./examples/05-healthcare-lattice/tools/generate_healthcare_workbook.py \
        --out ./examples/05-healthcare-lattice/CareOps_Workbook.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Facilities, domains — aligned to Healthcare Lattice README
# ---------------------------------------------------------------------------
FACILITIES = ["meridian-general", "meridian-community", "meridian-behavioral"]
DOMAINS = ["CLN", "REG", "OPS", "FIN"]

EPISODE_IDS = ["ep-hc-001", "ep-hc-002", "ep-hc-003", "ep-hc-004"]
DRIFT_IDS = ["DS-HC-001", "DS-HC-002", "DS-HC-003", "DS-HC-004"]
PATCH_IDS = ["Patch-HC-001", "Patch-HC-002", "Patch-HC-003", "Patch-HC-004"]


def _add_table(ws, name: str, rows: int) -> None:
    """Add a named Excel Table spanning all data in the worksheet."""
    max_col = get_column_letter(ws.max_column)
    ref = f"A1:{max_col}{rows + 1}"  # +1 for header
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _build_clinical_decisions(wb: Workbook) -> None:
    """ClinicalDecisions tab — 25 rows of clinical governance records."""
    ws = wb.create_sheet("ClinicalDecisions")
    headers = ["DecisionID", "Facility", "Domain", "Category",
               "Description", "Authority", "Outcome",
               "EpisodeRef", "Date"]
    ws.append(headers)

    rows = [
        ["CD-001", "meridian-general", "CLN", "Formulary", "Updated cardiac formulary to include new anticoagulant", "CMO", "Approved", "ep-hc-001", "2026-02-20"],
        ["CD-002", "meridian-community", "CLN", "Formulary", "Bridge protocol for transfer patients during formulary sync", "CPO", "Implemented", "ep-hc-001", "2026-02-20"],
        ["CD-003", "meridian-general", "CLN", "Protocol", "Sepsis bundle compliance — 3-hour window enforcement", "CMO", "Active", "", "2026-02-19"],
        ["CD-004", "meridian-behavioral", "CLN", "Protocol", "Updated restraint reduction protocol per CMS guidelines", "Medical Director", "Active", "", "2026-02-18"],
        ["CD-005", "meridian-general", "CLN", "Staffing", "Critical care float pool activation threshold lowered", "CNO", "Approved", "ep-hc-002", "2026-02-21"],
        ["CD-006", "meridian-community", "CLN", "Staffing", "Maternity unit staffing ratio exception during census spike", "CNO", "Temporary", "ep-hc-002", "2026-02-21"],
        ["CD-007", "meridian-general", "CLN", "Equipment", "Cardiac monitor calibration protocol revision", "Biomed Director", "Implemented", "ep-hc-004", "2026-02-22"],
        ["CD-008", "meridian-general", "CLN", "Protocol", "Fall prevention bundle update for geriatric unit", "CNO", "Active", "", "2026-02-17"],
        ["CD-009", "meridian-community", "CLN", "Protocol", "Pediatric dosing calculator update for new weight-based protocols", "Pharmacy Director", "Approved", "", "2026-02-17"],
        ["CD-010", "meridian-behavioral", "CLN", "Protocol", "De-escalation training requirement expanded to all clinical staff", "Medical Director", "Active", "", "2026-02-16"],
        ["CD-011", "meridian-general", "CLN", "Formulary", "Antibiotic stewardship — restrict broad-spectrum without ID consult", "Infectious Disease", "Active", "", "2026-02-15"],
        ["CD-012", "meridian-general", "CLN", "Equipment", "Infusion pump library update for new medication concentrations", "Pharmacy", "Deployed", "", "2026-02-15"],
        ["CD-013", "meridian-community", "CLN", "Protocol", "Post-surgical pain management protocol revision", "Anesthesia Chair", "Active", "", "2026-02-14"],
        ["CD-014", "meridian-general", "CLN", "Protocol", "Stroke code response time target reduced to 15 min", "Neuro Chief", "Active", "", "2026-02-14"],
        ["CD-015", "meridian-behavioral", "CLN", "Staffing", "1:1 observation staffing requirement for acute psychosis", "Medical Director", "Active", "", "2026-02-13"],
        ["CD-016", "meridian-general", "CLN", "Formulary", "Biosimilar substitution policy for oncology agents", "P&T Committee", "Approved", "", "2026-02-13"],
        ["CD-017", "meridian-community", "CLN", "Equipment", "Ventilator maintenance schedule revised per manufacturer update", "Respiratory Therapy", "Implemented", "", "2026-02-12"],
        ["CD-018", "meridian-general", "CLN", "Protocol", "Blood product transfusion consent process update", "CMO", "Active", "", "2026-02-12"],
        ["CD-019", "meridian-general", "CLN", "Protocol", "ICU sedation vacation protocol standardization", "ICU Director", "Active", "", "2026-02-11"],
        ["CD-020", "meridian-community", "CLN", "Formulary", "Emergency department medication kit composition review", "ED Medical Director", "Approved", "", "2026-02-11"],
        ["CD-021", "meridian-behavioral", "CLN", "Protocol", "Suicide risk assessment frequency increase for first 72h", "Medical Director", "Active", "", "2026-02-10"],
        ["CD-022", "meridian-general", "CLN", "Equipment", "Point-of-care testing device validation — new troponin assay", "Lab Director", "Validated", "", "2026-02-10"],
        ["CD-023", "meridian-community", "CLN", "Protocol", "Labor and delivery emergency cesarean readiness drill", "OB Chief", "Scheduled", "", "2026-02-09"],
        ["CD-024", "meridian-general", "CLN", "Formulary", "Controlled substance prescribing limits per DEA guidance", "CMO", "Active", "", "2026-02-09"],
        ["CD-025", "meridian-general", "CLN", "Protocol", "Rapid response team activation criteria update", "Patient Safety Officer", "Active", "", "2026-02-08"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblClinicalDecisions", len(rows))


def _build_staffing_models(wb: Workbook) -> None:
    """StaffingModels tab — 25 rows of staffing model records."""
    ws = wb.create_sheet("StaffingModels")
    headers = ["ModelID", "Facility", "Unit", "RatioRequired",
               "RatioActual", "Status", "Trigger",
               "EpisodeRef", "Date"]
    ws.append(headers)

    rows = [
        ["SM-001", "meridian-general", "ICU", "1:2", "1:2", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-002", "meridian-general", "Med-Surg", "1:5", "1:6", "Non-Compliant", "Census spike", "ep-hc-002", "2026-02-21"],
        ["SM-003", "meridian-community", "Maternity", "1:3", "1:4", "Warning", "Float pool depleted", "ep-hc-002", "2026-02-21"],
        ["SM-004", "meridian-behavioral", "Acute Psych", "1:4", "1:6", "Non-Compliant", "Census spike 48 -> 56", "ep-hc-002", "2026-02-21"],
        ["SM-005", "meridian-general", "ED", "1:4", "1:4", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-006", "meridian-general", "Cardiac", "1:3", "1:3", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-007", "meridian-community", "Pediatrics", "1:4", "1:4", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-008", "meridian-community", "General Surgery", "1:5", "1:5", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-009", "meridian-behavioral", "Crisis Stabilization", "1:3", "1:3", "Compliant", "Baseline", "", "2026-02-20"],
        ["SM-010", "meridian-general", "Oncology", "1:4", "1:4", "Compliant", "Baseline", "", "2026-02-19"],
        ["SM-011", "meridian-general", "NICU", "1:2", "1:2", "Compliant", "Baseline", "", "2026-02-19"],
        ["SM-012", "meridian-general", "Step-Down", "1:4", "1:5", "Warning", "Admission surge", "", "2026-02-19"],
        ["SM-013", "meridian-community", "ED", "1:4", "1:5", "Warning", "Evening shift gap", "", "2026-02-19"],
        ["SM-014", "meridian-general", "OR", "1:1", "1:1", "Compliant", "Baseline", "", "2026-02-18"],
        ["SM-015", "meridian-behavioral", "Group Therapy", "1:8", "1:8", "Compliant", "Baseline", "", "2026-02-18"],
        ["SM-016", "meridian-general", "Telemetry", "1:4", "1:4", "Compliant", "Baseline", "", "2026-02-17"],
        ["SM-017", "meridian-community", "Rehab", "1:6", "1:6", "Compliant", "Baseline", "", "2026-02-17"],
        ["SM-018", "meridian-general", "Float Pool", "N/A", "12 FTEs", "Active", "System reserve", "", "2026-02-20"],
        ["SM-019", "meridian-general", "Float Pool", "N/A", "4 FTEs", "Depleted", "Behavioral surge", "ep-hc-002", "2026-02-21"],
        ["SM-020", "meridian-general", "PACU", "1:2", "1:2", "Compliant", "Baseline", "", "2026-02-16"],
        ["SM-021", "meridian-community", "L&D", "1:2", "1:2", "Compliant", "Baseline", "", "2026-02-16"],
        ["SM-022", "meridian-behavioral", "Detox", "1:4", "1:5", "Warning", "Weekend coverage gap", "", "2026-02-15"],
        ["SM-023", "meridian-general", "Burn Unit", "1:2", "1:2", "Compliant", "Baseline", "", "2026-02-15"],
        ["SM-024", "meridian-general", "Dialysis", "1:3", "1:3", "Compliant", "Baseline", "", "2026-02-14"],
        ["SM-025", "meridian-community", "Nursery", "1:4", "1:4", "Compliant", "Baseline", "", "2026-02-14"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblStaffingModels", len(rows))


def _build_formulary_changes(wb: Workbook) -> None:
    """FormularyChanges tab — 25 rows of formulary change records."""
    ws = wb.create_sheet("FormularyChanges")
    headers = ["ChangeID", "Facility", "Domain", "Drug",
               "Action", "Reason", "P_T_Approval",
               "EpisodeRef", "Date"]
    ws.append(headers)

    rows = [
        ["RX-001", "meridian-general", "CLN", "Apixaban (Eliquis)", "Added to formulary", "New anticoagulant protocol", "Yes", "ep-hc-001", "2026-02-20"],
        ["RX-002", "meridian-community", "CLN", "Apixaban (Eliquis)", "Emergency push sync", "Cross-facility formulary reconciliation", "Yes", "ep-hc-001", "2026-02-20"],
        ["RX-003", "meridian-general", "CLN", "Ceftriaxone", "Restricted — ID consult required", "Antibiotic stewardship", "Yes", "", "2026-02-19"],
        ["RX-004", "meridian-general", "CLN", "Remdesivir", "Removed from formulary", "Clinical trial ended", "Yes", "", "2026-02-18"],
        ["RX-005", "meridian-community", "CLN", "Oxytocin", "Concentration standardized", "Safety alert — dose confusion", "Yes", "", "2026-02-18"],
        ["RX-006", "meridian-behavioral", "CLN", "Olanzapine IM", "Added fast-track protocol", "Acute agitation management", "Yes", "", "2026-02-17"],
        ["RX-007", "meridian-general", "CLN", "Trastuzumab biosimilar", "Substitution approved", "Cost reduction — $4.2M/yr savings", "Yes", "", "2026-02-17"],
        ["RX-008", "meridian-general", "CLN", "Vancomycin", "Updated dosing protocol", "Revised AUC-based dosing per ASHP", "Yes", "", "2026-02-16"],
        ["RX-009", "meridian-community", "CLN", "Ketorolac", "Duration limit reduced to 48h", "Renal risk in elderly", "Yes", "", "2026-02-16"],
        ["RX-010", "meridian-general", "CLN", "Insulin glargine", "Biosimilar substitution", "Formulary cost optimization", "Yes", "", "2026-02-15"],
        ["RX-011", "meridian-behavioral", "CLN", "Buprenorphine", "Expanded access — all prescribers", "DEA guideline change", "Yes", "", "2026-02-15"],
        ["RX-012", "meridian-general", "CLN", "Heparin", "Concentration standardized", "ISMP safety recommendation", "Yes", "", "2026-02-14"],
        ["RX-013", "meridian-community", "CLN", "Acetaminophen IV", "Added to ED formulary", "Pain management pathway update", "Yes", "", "2026-02-14"],
        ["RX-014", "meridian-general", "CLN", "Norepinephrine", "Concentration change", "Standardization across ICU/ED", "Yes", "", "2026-02-13"],
        ["RX-015", "meridian-general", "CLN", "Propofol", "Usage tracking enhanced", "Diversion prevention", "Yes", "", "2026-02-13"],
        ["RX-016", "meridian-community", "CLN", "Epinephrine auto-injector", "Par level increased", "Allergy season preparation", "N/A", "", "2026-02-12"],
        ["RX-017", "meridian-behavioral", "CLN", "Naloxone", "Standing order expanded", "Opioid response protocol", "Yes", "", "2026-02-12"],
        ["RX-018", "meridian-general", "CLN", "Dexmedetomidine", "Added ICU sedation protocol", "Sedation vacation compliance", "Yes", "", "2026-02-11"],
        ["RX-019", "meridian-general", "CLN", "TPA (Alteplase)", "Stroke code protocol update", "Door-to-needle time improvement", "Yes", "", "2026-02-11"],
        ["RX-020", "meridian-community", "CLN", "Magnesium sulfate", "Dosing protocol standardized", "Preeclampsia pathway update", "Yes", "", "2026-02-10"],
        ["RX-021", "meridian-general", "CLN", "Fentanyl patch", "Prescriber education required", "Safety alert — naive patients", "Yes", "", "2026-02-10"],
        ["RX-022", "meridian-behavioral", "CLN", "Clozapine", "Monitoring frequency updated", "REMS compliance", "Yes", "", "2026-02-09"],
        ["RX-023", "meridian-general", "CLN", "Nitroglycerin", "IV formulation switch", "Supply chain availability", "Yes", "", "2026-02-09"],
        ["RX-024", "meridian-community", "CLN", "Ondansetron", "Pediatric dosing update", "Weight-based recalculation", "Yes", "", "2026-02-08"],
        ["RX-025", "meridian-general", "CLN", "Hydromorphone", "Smart pump limit tightened", "High-alert medication review", "Yes", "", "2026-02-08"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblFormularyChanges", len(rows))


def _build_billing_rules(wb: Workbook) -> None:
    """BillingRules tab — 25 rows of billing/coding rule records."""
    ws = wb.create_sheet("BillingRules")
    headers = ["RuleID", "Facility", "Domain", "Category",
               "Description", "CMSRule", "Status",
               "EpisodeRef", "Date"]
    ws.append(headers)

    rows = [
        ["BR-001", "meridian-general", "FIN", "Observation", "Observation-to-inpatient reclassification criteria", "CMS-2024-0019", "Stale — 3 weeks", "ep-hc-003", "2026-02-21"],
        ["BR-002", "meridian-general", "FIN", "Coding", "Cardiac DRG coding update for new procedures", "CMS-2024-0022", "Current", "", "2026-02-20"],
        ["BR-003", "meridian-community", "FIN", "Coding", "Maternity bundled payment coding rules", "CMS-2024-0015", "Current", "", "2026-02-20"],
        ["BR-004", "meridian-general", "FIN", "Denial", "Prior authorization workflow for high-cost imaging", "Payer-specific", "Active", "", "2026-02-19"],
        ["BR-005", "meridian-general", "FIN", "Compliance", "RAC audit response protocol", "CMS-RAC", "Active", "", "2026-02-19"],
        ["BR-006", "meridian-community", "FIN", "Observation", "Two-midnight rule compliance tracking", "CMS-1599-F", "Current", "", "2026-02-18"],
        ["BR-007", "meridian-general", "FIN", "Coding", "Sepsis coding — POA indicator rules", "CMS-ICD10-2026", "Current", "", "2026-02-18"],
        ["BR-008", "meridian-behavioral", "FIN", "Coding", "Behavioral health partial hospitalization billing", "CMS-BH-2025", "Current", "", "2026-02-17"],
        ["BR-009", "meridian-general", "FIN", "Denial", "Top 10 denial codes — quarterly review", "Internal", "Active", "ep-hc-003", "2026-02-21"],
        ["BR-010", "meridian-general", "FIN", "Revenue", "Charge capture audit — surgical services", "Internal", "Active", "", "2026-02-17"],
        ["BR-011", "meridian-community", "FIN", "Coding", "ED E/M level documentation requirements", "CMS-2024-0018", "Current", "", "2026-02-16"],
        ["BR-012", "meridian-general", "FIN", "Compliance", "Stark Law compliance — physician referral tracking", "42 CFR 411", "Active", "", "2026-02-16"],
        ["BR-013", "meridian-general", "FIN", "Revenue", "Self-pay financial counseling workflow", "Internal", "Active", "", "2026-02-15"],
        ["BR-014", "meridian-behavioral", "FIN", "Coding", "Substance abuse treatment coding — ASAM criteria", "CMS-SUD-2025", "Current", "", "2026-02-15"],
        ["BR-015", "meridian-general", "FIN", "Denial", "Payer contract escalation — 30-day AR threshold", "Payer-specific", "Active", "", "2026-02-14"],
        ["BR-016", "meridian-community", "FIN", "Revenue", "Outpatient surgery center fee schedule update", "CMS-OPPS-2026", "Current", "", "2026-02-14"],
        ["BR-017", "meridian-general", "FIN", "Compliance", "False Claims Act awareness — annual attestation", "31 USC 3729", "Active", "", "2026-02-13"],
        ["BR-018", "meridian-general", "FIN", "Coding", "Modifier 25 documentation standard", "CMS-2024-0020", "Current", "", "2026-02-13"],
        ["BR-019", "meridian-community", "FIN", "Denial", "Timely filing limit tracker — payer matrix", "Payer-specific", "Active", "", "2026-02-12"],
        ["BR-020", "meridian-general", "FIN", "Revenue", "Bad debt write-off approval workflow", "Internal", "Active", "", "2026-02-12"],
        ["BR-021", "meridian-behavioral", "FIN", "Compliance", "Medicaid billing — concurrent review documentation", "State Medicaid", "Active", "", "2026-02-11"],
        ["BR-022", "meridian-general", "FIN", "Coding", "Transplant DRG coding update", "CMS-2024-0025", "Current", "", "2026-02-11"],
        ["BR-023", "meridian-community", "FIN", "Revenue", "Charity care application processing — 30-day SLA", "Internal", "Active", "", "2026-02-10"],
        ["BR-024", "meridian-general", "FIN", "Compliance", "Anti-Kickback Statute compliance — vendor review", "42 USC 1320a-7b", "Active", "", "2026-02-10"],
        ["BR-025", "meridian-general", "FIN", "Denial", "Clinical documentation improvement (CDI) query response rate", "Internal", "Active", "", "2026-02-09"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblBillingRules", len(rows))


def _build_assumptions(wb: Workbook) -> None:
    """Assumptions tab — 25 rows of institutional assumptions with TTL."""
    ws = wb.create_sheet("Assumptions")
    headers = ["AssumptionID", "Facility", "Domain", "Assumption",
               "Confidence", "TTL_Days", "LastValidated", "Source",
               "RiskIfWrong", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["A-001", "All", "CLN", "Formulary is synchronized across all facilities within 4 hours of change", "0.50", "7", "2026-02-20", "S-RX-001", "Dispensing conflict at transfer", "ep-hc-001"],
        ["A-002", "All", "OPS", "Float pool capacity sufficient for 15% census variance", "0.45", "7", "2026-02-21", "S-HR-001", "Staffing ratio non-compliance", "ep-hc-002"],
        ["A-003", "meridian-general", "FIN", "Billing templates current within 2 weeks of CMS updates", "0.40", "14", "2026-02-21", "S-REV-001", "Upcoding exposure; claim denials", "ep-hc-003"],
        ["A-004", "meridian-general", "OPS", "Equipment calibration handoff is complete between shifts", "0.55", "7", "2026-02-22", "S-BIO-001", "False alarm suppression; delayed detection", "ep-hc-004"],
        ["A-005", "All", "REG", "CMS Conditions of Participation unchanged this quarter", "0.85", "90", "2026-02-15", "S-INC-001", "Survey readiness gap", ""],
        ["A-006", "All", "REG", "Joint Commission standards stable for accreditation cycle", "0.80", "180", "2026-01-15", "S-INC-001", "Accreditation risk", ""],
        ["A-007", "All", "CLN", "Clinical pathways reflect current evidence-based guidelines", "0.75", "30", "2026-02-10", "S-EHR-001", "Suboptimal care delivery", ""],
        ["A-008", "meridian-general", "CLN", "Trauma Level II capabilities maintained 24/7", "0.90", "7", "2026-02-19", "S-EHR-001", "Diversion; patient safety", ""],
        ["A-009", "meridian-community", "CLN", "Maternity unit staffing meets AWHONN guidelines", "0.70", "7", "2026-02-19", "S-HR-001", "Patient safety; regulatory citation", ""],
        ["A-010", "meridian-behavioral", "CLN", "Restraint use declining per de-escalation initiative", "0.65", "14", "2026-02-18", "S-INC-001", "CMS deficiency", ""],
        ["A-011", "All", "FIN", "Payer contract terms unchanged through Q1 2026", "0.80", "60", "2026-01-05", "S-REV-001", "Revenue variance", ""],
        ["A-012", "All", "OPS", "Supply chain lead times stable for critical items", "0.60", "14", "2026-02-15", "S-SUP-001", "Stockout; care pathway disruption", "ep-hc-004"],
        ["A-013", "All", "REG", "HIPAA enforcement posture unchanged", "0.85", "90", "2026-01-10", "S-INC-001", "PHI breach penalty", ""],
        ["A-014", "meridian-general", "FIN", "RAC audit findings remain below threshold", "0.70", "30", "2026-02-01", "S-REV-001", "Repayment demand", ""],
        ["A-015", "All", "OPS", "EHR uptime >= 99.5% (Epic SLA)", "0.90", "30", "2026-02-15", "S-EHR-001", "Clinical workflow disruption", ""],
        ["A-016", "meridian-general", "CLN", "Blood bank inventory sufficient for trauma volume", "0.80", "7", "2026-02-19", "S-SUP-001", "Transfusion delay", ""],
        ["A-017", "All", "FIN", "Charity care policy compliant with 501(c)(3) requirements", "0.85", "90", "2026-01-20", "S-REV-001", "Tax-exempt status risk", ""],
        ["A-018", "meridian-behavioral", "OPS", "Crisis stabilization bed capacity adequate for community need", "0.55", "14", "2026-02-18", "S-HR-001", "Diversion; community impact", "ep-hc-002"],
        ["A-019", "All", "REG", "State licensure requirements unchanged", "0.90", "180", "2026-01-01", "S-INC-001", "Operational authority risk", ""],
        ["A-020", "meridian-general", "OPS", "Biomed maintenance schedule current for all critical equipment", "0.65", "7", "2026-02-22", "S-BIO-001", "Equipment failure; patient safety", "ep-hc-004"],
        ["A-021", "All", "FIN", "Medicare reimbursement rates stable for FY2026", "0.85", "365", "2026-01-01", "S-REV-001", "Budget variance", ""],
        ["A-022", "meridian-community", "OPS", "Ambulance response time to Community campus < 8 min", "0.75", "30", "2026-02-10", "S-INC-001", "Transfer delay", ""],
        ["A-023", "All", "CLN", "Medication error rate below industry benchmark", "0.70", "30", "2026-02-15", "S-INC-001", "Patient safety event", ""],
        ["A-024", "meridian-general", "REG", "DEA controlled substance audit current", "0.80", "90", "2026-01-15", "S-INC-001", "DEA enforcement action", ""],
        ["A-025", "All", "OPS", "Cross-facility patient transfer protocol effective", "0.60", "14", "2026-02-20", "S-EHR-001", "Care continuity gap", "ep-hc-001"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblAssumptions", len(rows))


def _build_drift_signals(wb: Workbook) -> None:
    """DriftSignals tab — 25 rows including the 4 scenario signals."""
    ws = wb.create_sheet("DriftSignals")
    headers = ["DriftID", "Facility", "Domain", "Category", "Severity",
               "ClaimsAffected", "EvidenceCount", "DetectedAt",
               "Status", "PatchRef", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["DS-HC-001", "meridian-general", "CLN/REG/FIN", "formulary_mismatch", "RED", "CLN-001,CLN-002,REG-001", "14", "2026-02-20T18:00:00Z", "Patched", "Patch-HC-001", "ep-hc-001"],
        ["DS-HC-002", "All", "OPS/CLN/REG", "staffing_cascade", "RED", "OPS-001,OPS-002,CLN-003,REG-002", "22", "2026-02-21T06:00:00Z", "Patched", "Patch-HC-002", "ep-hc-002"],
        ["DS-HC-003", "meridian-general", "FIN/REG/CLN", "billing_classification_drift", "RED", "FIN-002,REG-003,CLN-004", "18", "2026-02-21T14:00:00Z", "Patched", "Patch-HC-003", "ep-hc-003"],
        ["DS-HC-004", "meridian-general", "OPS/CLN", "equipment_verification_gap", "YELLOW", "OPS-003,CLN-005", "8", "2026-02-22T03:00:00Z", "Patched", "Patch-HC-004", "ep-hc-004"],
        ["DS-HC-005", "meridian-community", "CLN", "antibiotic_stewardship_drift", "YELLOW", "CLN-006", "5", "2026-02-19T10:00:00Z", "Active", "", ""],
        ["DS-HC-006", "meridian-general", "REG", "consent_documentation_gap", "YELLOW", "REG-004", "4", "2026-02-18T14:00:00Z", "Active", "", ""],
        ["DS-HC-007", "meridian-behavioral", "CLN", "restraint_protocol_variance", "YELLOW", "CLN-007", "6", "2026-02-18T08:00:00Z", "Monitoring", "", ""],
        ["DS-HC-008", "meridian-general", "FIN", "charge_capture_leakage", "YELLOW", "FIN-003", "8", "2026-02-17T16:00:00Z", "Active", "", ""],
        ["DS-HC-009", "All", "OPS", "ehr_downtime_frequency", "GREEN", "OPS-004", "3", "2026-02-17T10:00:00Z", "Monitoring", "", ""],
        ["DS-HC-010", "meridian-general", "CLN", "sepsis_bundle_compliance_drop", "YELLOW", "CLN-008", "7", "2026-02-16T22:00:00Z", "Active", "", ""],
        ["DS-HC-011", "meridian-community", "REG", "survey_readiness_gap", "YELLOW", "REG-005", "5", "2026-02-16T14:00:00Z", "Active", "", ""],
        ["DS-HC-012", "meridian-general", "FIN", "prior_auth_backlog_growing", "YELLOW", "FIN-004", "6", "2026-02-15T11:00:00Z", "Active", "", ""],
        ["DS-HC-013", "meridian-general", "CLN", "blood_product_waste_elevated", "GREEN", "CLN-009", "4", "2026-02-15T08:00:00Z", "Monitoring", "", ""],
        ["DS-HC-014", "meridian-behavioral", "OPS", "bed_capacity_approaching_limit", "YELLOW", "OPS-005", "5", "2026-02-14T16:00:00Z", "Active", "", ""],
        ["DS-HC-015", "meridian-general", "REG", "incident_reporting_lag", "GREEN", "REG-006", "3", "2026-02-14T10:00:00Z", "Monitoring", "", ""],
        ["DS-HC-016", "meridian-community", "FIN", "denial_rate_trending_up", "YELLOW", "FIN-005", "7", "2026-02-13T14:00:00Z", "Active", "", ""],
        ["DS-HC-017", "meridian-general", "OPS", "supply_chain_lead_time_increase", "YELLOW", "OPS-006", "5", "2026-02-13T09:00:00Z", "Active", "", ""],
        ["DS-HC-018", "meridian-general", "CLN", "fall_rate_above_benchmark", "YELLOW", "CLN-010", "6", "2026-02-12T20:00:00Z", "Active", "", ""],
        ["DS-HC-019", "All", "OPS", "float_pool_utilization_high", "YELLOW", "OPS-007", "8", "2026-02-12T12:00:00Z", "Monitoring", "", ""],
        ["DS-HC-020", "meridian-general", "FIN", "ar_days_above_target", "GREEN", "FIN-006", "4", "2026-02-11T15:00:00Z", "Monitoring", "", ""],
        ["DS-HC-021", "meridian-community", "CLN", "medication_error_near_miss", "YELLOW", "CLN-011", "3", "2026-02-11T08:00:00Z", "Active", "", ""],
        ["DS-HC-022", "meridian-behavioral", "REG", "hipaa_training_overdue", "YELLOW", "REG-007", "4", "2026-02-10T14:00:00Z", "Active", "", ""],
        ["DS-HC-023", "meridian-general", "OPS", "or_utilization_below_target", "GREEN", "OPS-008", "3", "2026-02-10T10:00:00Z", "Monitoring", "", ""],
        ["DS-HC-024", "meridian-community", "CLN", "pain_management_protocol_lag", "GREEN", "CLN-012", "3", "2026-02-09T16:00:00Z", "Monitoring", "", ""],
        ["DS-HC-025", "meridian-general", "REG", "controlled_substance_audit_due", "YELLOW", "REG-008", "4", "2026-02-09T09:00:00Z", "Active", "", ""],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblDriftSignals", len(rows))


def _build_patch_plans(wb: Workbook) -> None:
    """PatchPlans tab — 25 rows of patch plan records."""
    ws = wb.create_sheet("PatchPlans")
    headers = ["PatchID", "Title", "DriftRef", "Severity", "Domains",
               "SelectedOption", "Steps", "Owner", "Status",
               "ClosureConditions", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["Patch-HC-001", "Cross-facility formulary sync", "DS-HC-001", "RED", "CLN/REG/FIN", "C", "3", "chief-pharmacy-officer", "Applied", "CLN-001>=0.85; all facilities synced; no safety events", "ep-hc-001"],
        ["Patch-HC-002", "Staffing allocation model", "DS-HC-002", "RED", "OPS/CLN/REG", "A", "3", "chief-nursing-officer", "Applied", "All ratios compliant; float pool restored; no citations", "ep-hc-002"],
        ["Patch-HC-003", "Billing/coding drift fix", "DS-HC-003", "RED", "FIN/OPS/REG", "A", "3", "revenue-cycle", "Applied", "Denial rate<=baseline; FIN>=0.85; no RED FIN/REG", "ep-hc-003"],
        ["Patch-HC-004", "Equipment calibration gap", "DS-HC-004", "YELLOW", "CLN/OPS", "A", "2", "clinical-lead", "Applied", "Zero stockouts 14d; OPS>=0.80", "ep-hc-004"],
        ["Patch-HC-005", "Antibiotic stewardship enforcement", "DS-HC-005", "YELLOW", "CLN", "-", "2", "infectious-disease", "Planned", "Stewardship compliance >= 90%", ""],
        ["Patch-HC-006", "Consent documentation improvement", "DS-HC-006", "YELLOW", "REG", "-", "1", "compliance", "Planned", "Consent completion rate >= 98%", ""],
        ["Patch-HC-007", "Restraint protocol alignment", "DS-HC-007", "YELLOW", "CLN", "-", "2", "medical-director", "Monitoring", "Restraint rate declining trend", ""],
        ["Patch-HC-008", "Charge capture optimization", "DS-HC-008", "YELLOW", "FIN", "-", "2", "revenue-cycle", "In Progress", "Leakage rate < 2%", ""],
        ["Patch-HC-009", "Sepsis bundle compliance", "DS-HC-010", "YELLOW", "CLN", "-", "2", "quality", "In Progress", "Bundle compliance >= 85%", ""],
        ["Patch-HC-010", "Survey readiness remediation", "DS-HC-011", "YELLOW", "REG", "-", "3", "compliance", "Planned", "All survey deficiencies addressed", ""],
        ["Patch-HC-011", "Prior auth backlog reduction", "DS-HC-012", "YELLOW", "FIN", "-", "2", "revenue-cycle", "In Progress", "Backlog < 48h", ""],
        ["Patch-HC-012", "Behavioral capacity expansion", "DS-HC-014", "YELLOW", "OPS", "-", "2", "ops-director", "Planned", "Capacity margin >= 15%", ""],
        ["Patch-HC-013", "Denial rate intervention", "DS-HC-016", "YELLOW", "FIN", "-", "2", "revenue-cycle", "In Progress", "Denial rate below Q4 baseline", ""],
        ["Patch-HC-014", "Supply chain buffer increase", "DS-HC-017", "YELLOW", "OPS", "-", "2", "supply-chain", "Planned", "Critical item buffer >= 14 days", ""],
        ["Patch-HC-015", "Fall prevention bundle", "DS-HC-018", "YELLOW", "CLN", "-", "2", "patient-safety", "In Progress", "Fall rate below benchmark", ""],
        ["Patch-HC-016", "Float pool capacity planning", "DS-HC-019", "YELLOW", "OPS", "-", "1", "cno", "Monitoring", "Float utilization < 80%", ""],
        ["Patch-HC-017", "Medication error prevention", "DS-HC-021", "YELLOW", "CLN", "-", "2", "pharmacy", "Planned", "Zero preventable errors 30d", ""],
        ["Patch-HC-018", "HIPAA training completion", "DS-HC-022", "YELLOW", "REG", "-", "1", "compliance", "In Progress", "100% completion within 30d", ""],
        ["Patch-HC-019", "OR utilization optimization", "DS-HC-023", "GREEN", "OPS", "-", "2", "surgical-services", "Monitoring", "Utilization >= 75%", ""],
        ["Patch-HC-020", "Pain management protocol update", "DS-HC-024", "GREEN", "CLN", "-", "1", "quality", "Monitoring", "Protocol current", ""],
        ["Patch-HC-021", "Controlled substance audit", "DS-HC-025", "YELLOW", "REG", "-", "1", "compliance", "Planned", "Audit current; zero findings", ""],
        ["Patch-HC-022", "Blood product waste reduction", "DS-HC-013", "GREEN", "CLN", "-", "1", "transfusion-committee", "Monitoring", "Waste rate < 5%", ""],
        ["Patch-HC-023", "Incident reporting timeliness", "DS-HC-015", "GREEN", "REG", "-", "1", "patient-safety", "Monitoring", "Reporting lag < 24h", ""],
        ["Patch-HC-024", "AR days reduction", "DS-HC-020", "GREEN", "FIN", "-", "2", "revenue-cycle", "Monitoring", "AR days < target", ""],
        ["Patch-HC-025", "EHR uptime improvement", "DS-HC-009", "GREEN", "OPS", "-", "1", "health-it", "Monitoring", "Uptime >= 99.5%", ""],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblPatchPlans", len(rows))


def _build_canon_rules(wb: Workbook) -> None:
    """CanonRules tab — 25 rows of institutional governance rules."""
    ws = wb.create_sheet("CanonRules")
    headers = ["RuleID", "Facility", "Domain", "Rule", "Tier",
               "Enforcement", "Source", "LastReviewed"]
    ws.append(headers)

    rows = [
        ["CR-001", "All", "CLN", "Formulary changes must sync to all facilities within 4 hours", "0", "Blocking", "S-RX-001", "2026-02-20"],
        ["CR-002", "All", "CLN", "Patient transfer requires medication reconciliation at both ends", "0", "Blocking", "S-EHR-001", "2026-02-20"],
        ["CR-003", "All", "OPS", "Staffing ratios must meet CMS and state minimum requirements", "0", "Blocking", "S-HR-001", "2026-02-21"],
        ["CR-004", "All", "OPS", "Float pool depletion triggers immediate leadership escalation", "0", "Blocking", "S-HR-001", "2026-02-21"],
        ["CR-005", "All", "FIN", "Billing templates must be current within 14 days of CMS updates", "0", "Blocking", "S-REV-001", "2026-02-21"],
        ["CR-006", "All", "FIN", "Revenue anomalies > 20% trigger automatic audit", "0", "Blocking", "S-REV-001", "2026-02-21"],
        ["CR-007", "All", "OPS", "Equipment calibration must be verified at every shift handoff", "1", "Gating", "S-BIO-001", "2026-02-22"],
        ["CR-008", "All", "REG", "CMS Conditions of Participation compliance continuous", "0", "Blocking", "S-INC-001", "2026-02-15"],
        ["CR-009", "All", "REG", "Joint Commission standards readiness maintained at all times", "0", "Blocking", "S-INC-001", "2026-02-15"],
        ["CR-010", "All", "CLN", "Antibiotic stewardship requires ID consult for restricted agents", "1", "Gating", "S-RX-001", "2026-02-19"],
        ["CR-011", "All", "CLN", "High-alert medication changes require dual verification", "0", "Blocking", "S-RX-001", "2026-02-15"],
        ["CR-012", "All", "REG", "HIPAA training must be current for all workforce members", "0", "Blocking", "S-INC-001", "2026-02-10"],
        ["CR-013", "All", "FIN", "Claim submission within timely filing limits per payer contract", "1", "Gating", "S-REV-001", "2026-02-14"],
        ["CR-014", "All", "OPS", "Critical supply par levels reviewed weekly", "1", "Gating", "S-SUP-001", "2026-02-15"],
        ["CR-015", "All", "CLN", "Restraint use requires physician order and Q15 monitoring", "0", "Blocking", "S-INC-001", "2026-02-17"],
        ["CR-016", "All", "FIN", "Controlled substance prescribing within DEA guidelines", "0", "Blocking", "S-RX-001", "2026-02-09"],
        ["CR-017", "All", "CLN", "Sepsis bundle must be initiated within 3 hours of identification", "0", "Blocking", "S-EHR-001", "2026-02-16"],
        ["CR-018", "All", "OPS", "Patient transport protocols require clinical handoff documentation", "1", "Gating", "S-EHR-001", "2026-02-14"],
        ["CR-019", "All", "REG", "Incident reports must be filed within 24 hours of event", "1", "Gating", "S-INC-001", "2026-02-14"],
        ["CR-020", "All", "CLN", "Blood product administration requires two-nurse verification", "0", "Blocking", "S-EHR-001", "2026-02-12"],
        ["CR-021", "All", "FIN", "Charity care eligibility screening within 5 business days", "2", "Advisory", "S-REV-001", "2026-02-10"],
        ["CR-022", "All", "OPS", "Emergency backup power tested monthly", "1", "Gating", "S-BIO-001", "2026-02-08"],
        ["CR-023", "All", "REG", "Privacy breach notification within 60 days per HIPAA", "0", "Blocking", "S-INC-001", "2026-02-08"],
        ["CR-024", "All", "CLN", "Code Blue response time < 3 minutes from activation", "0", "Blocking", "S-INC-001", "2026-02-07"],
        ["CR-025", "All", "OPS", "Cross-facility patient transfer protocol includes formulary check", "1", "Gating", "S-EHR-001", "2026-02-20"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblCanonRules", len(rows))


def _build_prompts(wb: Workbook) -> None:
    """PROMPTS tab — LLM interaction surface."""
    ws = wb.create_sheet("PROMPTS")

    ws["A1"] = "What would you like to do today?"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Options:"
    ws["A4"] = "(1) WHY did we change X?"
    ws["A5"] = "(2) WHAT drifted?"
    ws["A6"] = "(3) Generate patch plan"
    ws["A7"] = "(4) Audit formulary / staffing / billing loops"

    ws["A9"] = "System Prompt (for LLM integration):"
    ws["A9"].font = Font(bold=True)

    system_prompt = (
        "You are a CareOps governance assistant for Meridian Health Partners (fictional). "
        "Read ALL tabs in this workbook: ClinicalDecisions, StaffingModels, FormularyChanges, "
        "BillingRules, Assumptions, DriftSignals, PatchPlans, CanonRules. Cross-reference episode IDs "
        "(ep-hc-XXX), drift IDs (DS-HC-XXX), and patch IDs (Patch-HC-XXX) across tabs. "
        "When the user asks a question:\n"
        "1. Identify which tabs contain relevant data\n"
        "2. Trace the decision chain: ClinicalDecision/FormularyChange -> DriftSignal -> PatchPlan -> Episode\n"
        "3. Check CanonRules for any governance violations\n"
        "4. Check Assumptions for any expired or low-confidence beliefs\n"
        "5. Respond with: (a) what happened, (b) why it matters, (c) what to do next\n"
        "6. Link to specific row IDs so the user can verify\n\n"
        "Facilities: Meridian General (450 beds), Meridian Community (180 beds), Meridian Behavioral (60 beds)\n"
        "Domains: CLN (Clinical), REG (Regulatory), OPS (Operational), FIN (Financial)\n"
        "Guardrails: Fictional data only. No real patients, clinicians, or facilities. No clinical advice."
    )
    ws["A10"] = system_prompt

    # Set column width for readability
    ws.column_dimensions["A"].width = 100


def generate_workbook(output_path: Path) -> None:
    """Generate the complete CareOps Pack workbook."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _build_clinical_decisions(wb)
    _build_staffing_models(wb)
    _build_formulary_changes(wb)
    _build_billing_rules(wb)
    _build_assumptions(wb)
    _build_drift_signals(wb)
    _build_patch_plans(wb)
    _build_canon_rules(wb)
    _build_prompts(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"CareOps Workbook generated: {output_path}")
    print(f"  Tabs: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
    print(f"  Rows: {25 * 8 + 10} (25 data rows x 8 tables + PROMPTS)")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate CareOps Pack Excel workbook",
    )
    parser.add_argument(
        "--out",
        type=str,
        default="./CareOps_Workbook.xlsx",
        help="Output path for the workbook (default: ./CareOps_Workbook.xlsx)",
    )
    args = parser.parse_args()
    generate_workbook(Path(args.out))


if __name__ == "__main__":
    main()
