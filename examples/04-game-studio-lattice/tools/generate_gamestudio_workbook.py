#!/usr/bin/env python3
"""Generate GameOps Pack Excel workbook for the Game Studio Lattice example.

Creates a multi-tab workbook with synthetic data aligned to the Nexus Interactive
(fictional) AAA publisher scenario. Each tab is a named Excel Table.

Requires: openpyxl (pip install openpyxl)

Usage:
    python ./examples/04-game-studio-lattice/tools/generate_gamestudio_workbook.py
    python ./examples/04-game-studio-lattice/tools/generate_gamestudio_workbook.py \
        --out ./examples/04-game-studio-lattice/GameOps_Workbook.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Studios, titles, domains — aligned to Game Studio Lattice README
# ---------------------------------------------------------------------------
STUDIOS = ["tokyo", "montreal", "bucharest", "singapore"]
TITLES = ["RONIN", "VANGUARD", "SIGNAL"]
DOMAINS = ["CRE", "REG", "PLT", "MON", "OPS", "DAT"]

EPISODE_IDS = ["ep-gs-001", "ep-gs-002", "ep-gs-003", "ep-gs-004"]
DRIFT_IDS = ["DS-GS-001", "DS-GS-002", "DS-GS-003", "DS-GS-004"]
PATCH_IDS = ["Patch-GS-001", "Patch-GS-002", "Patch-GS-003", "Patch-GS-004"]


def _add_table(ws, name: str, rows: int) -> None:
    """Add a named Excel Table spanning all data in the worksheet."""
    max_col = get_column_letter(ws.max_column)
    ref = f"A1:{max_col}{rows + 1}"  # +1 for header
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _build_balance_changes(wb: Workbook) -> None:
    """BalanceChanges tab — 25 rows of synthetic balance tuning records."""
    ws = wb.create_sheet("BalanceChanges")
    headers = ["ChangeID", "Title", "Studio", "Domain", "Category",
               "Parameter", "OldValue", "NewValue", "Reason",
               "EpisodeRef", "ApprovedBy", "Date"]
    ws.append(headers)

    rows = [
        ["BC-001", "RONIN", "tokyo", "CRE", "Combat", "dismemberment_vfx_enabled", "true", "false", "PEGI16 compliance", "ep-gs-001", "creative_director", "2026-02-19"],
        ["BC-002", "RONIN", "tokyo", "CRE", "Combat", "blood_particle_count", "200", "80", "Rating envelope", "ep-gs-001", "art_lead", "2026-02-19"],
        ["BC-003", "VANGUARD", "singapore", "MON", "Economy", "founders_cache_price_usd", "9.99", "4.99", "Monetization review", "ep-gs-002", "live_ops_lead", "2026-02-19"],
        ["BC-004", "VANGUARD", "singapore", "MON", "Economy", "gacha_pity_timer", "80", "60", "Player retention", "ep-gs-002", "economy_designer", "2026-02-19"],
        ["BC-005", "SIGNAL", "singapore", "MON", "Economy", "daily_gem_cap", "500", "300", "Spending control", "ep-gs-002", "economy_designer", "2026-02-19"],
        ["BC-006", "RONIN", "tokyo", "CRE", "Difficulty", "boss_hp_multiplier", "1.0", "1.2", "Challenge tuning", "", "combat_designer", "2026-02-18"],
        ["BC-007", "VANGUARD", "tokyo", "CRE", "Weapons", "ar_damage_falloff", "0.85", "0.78", "Competitive balance", "", "weapons_designer", "2026-02-18"],
        ["BC-008", "VANGUARD", "tokyo", "CRE", "Weapons", "smg_fire_rate", "720", "680", "TTK alignment", "", "weapons_designer", "2026-02-18"],
        ["BC-009", "SIGNAL", "montreal", "CRE", "Events", "event_reward_multiplier", "1.5", "1.2", "Economy normalization", "", "live_ops_lead", "2026-02-17"],
        ["BC-010", "RONIN", "montreal", "CRE", "Narrative", "dlc_cutscene_count", "8", "6", "Scope reduction", "ep-gs-001", "narrative_director", "2026-02-19"],
        ["BC-011", "VANGUARD", "singapore", "OPS", "Matchmaking", "sbmm_bracket_width", "150", "100", "Skill compression", "", "systems_designer", "2026-02-17"],
        ["BC-012", "VANGUARD", "singapore", "OPS", "Matchmaking", "max_ping_threshold_ms", "80", "120", "APAC region coverage", "", "network_eng", "2026-02-17"],
        ["BC-013", "RONIN", "bucharest", "PLT", "Performance", "target_fps_switch", "30", "30", "Unchanged after review", "", "perf_lead", "2026-02-16"],
        ["BC-014", "VANGUARD", "bucharest", "PLT", "Compliance", "cn_playtime_warning_min", "60", "45", "NPPA update", "ep-gs-004", "compliance_eng", "2026-02-19"],
        ["BC-015", "SIGNAL", "bucharest", "PLT", "Compliance", "age_gate_kr_enabled", "false", "true", "KR regulation", "ep-gs-004", "compliance_eng", "2026-02-19"],
        ["BC-016", "RONIN", "tokyo", "CRE", "Audio", "gore_sfx_volume_db", "-3", "-12", "Rating alignment", "ep-gs-001", "audio_director", "2026-02-19"],
        ["BC-017", "VANGUARD", "singapore", "MON", "Store", "battle_pass_tier_count", "100", "80", "Season length", "", "monetization_lead", "2026-02-16"],
        ["BC-018", "SIGNAL", "singapore", "MON", "Store", "gacha_banner_duration_hrs", "72", "48", "Rotation speed", "", "monetization_lead", "2026-02-16"],
        ["BC-019", "RONIN", "montreal", "DAT", "Privacy", "telemetry_opt_in_default", "true", "false", "GDPR alignment", "", "privacy_eng", "2026-02-15"],
        ["BC-020", "VANGUARD", "montreal", "DAT", "Privacy", "replay_data_retention_days", "90", "30", "CCPA request", "", "privacy_eng", "2026-02-15"],
        ["BC-021", "RONIN", "tokyo", "CRE", "Combat", "finisher_camera_shake", "1.0", "0.5", "Accessibility", "", "ux_designer", "2026-02-14"],
        ["BC-022", "VANGUARD", "tokyo", "OPS", "Server", "tick_rate_hz", "60", "64", "Competitive standard", "", "server_eng", "2026-02-14"],
        ["BC-023", "SIGNAL", "montreal", "OPS", "Events", "maintenance_window_utc", "06:00", "04:00", "APAC overlap", "", "live_ops_lead", "2026-02-13"],
        ["BC-024", "RONIN", "bucharest", "REG", "Ratings", "content_descriptor_violence", "moderate", "strong", "Pre-check update", "ep-gs-001", "qa_lead", "2026-02-19"],
        ["BC-025", "VANGUARD", "bucharest", "REG", "Ratings", "lootbox_disclosure_visible", "false", "true", "BE/NL compliance", "ep-gs-002", "qa_lead", "2026-02-19"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblBalanceChanges", len(rows))


def _build_economy_tuning(wb: Workbook) -> None:
    """EconomyTuning tab — 25 rows of economy parameter changes."""
    ws = wb.create_sheet("EconomyTuning")
    headers = ["TuneID", "Title", "Studio", "Parameter", "Category",
               "OldValue", "NewValue", "Impact", "EpisodeRef", "Date"]
    ws.append(headers)

    rows = [
        ["ET-001", "VANGUARD", "singapore", "premium_currency_earn_rate", "Earn", "50/hr", "35/hr", "Slower F2P progression", "ep-gs-002", "2026-02-19"],
        ["ET-002", "VANGUARD", "singapore", "battle_pass_xp_per_match", "Progression", "100", "120", "Faster BP completion", "", "2026-02-18"],
        ["ET-003", "SIGNAL", "singapore", "gacha_5star_rate", "Gacha", "0.6%", "0.8%", "Pity improvement", "ep-gs-002", "2026-02-19"],
        ["ET-004", "SIGNAL", "singapore", "stamina_refresh_cost", "Energy", "50 gems", "30 gems", "Retention boost", "", "2026-02-17"],
        ["ET-005", "VANGUARD", "singapore", "ranked_reward_multiplier", "Ranked", "1.5x", "2.0x", "Competitive incentive", "", "2026-02-17"],
        ["ET-006", "SIGNAL", "singapore", "daily_login_gem_bonus", "Login", "100", "150", "DAU retention", "", "2026-02-16"],
        ["ET-007", "VANGUARD", "singapore", "store_discount_max_pct", "Store", "30%", "25%", "Revenue protection", "", "2026-02-16"],
        ["ET-008", "SIGNAL", "singapore", "event_currency_exchange_rate", "Events", "10:1", "8:1", "Event value increase", "", "2026-02-15"],
        ["ET-009", "VANGUARD", "singapore", "founders_cache_item_count", "Bundle", "5", "8", "Value perception", "ep-gs-002", "2026-02-19"],
        ["ET-010", "SIGNAL", "singapore", "guild_donation_reward", "Social", "20 gems", "30 gems", "Guild engagement", "", "2026-02-15"],
        ["ET-011", "VANGUARD", "singapore", "weapon_skin_price_usd", "Cosmetic", "19.99", "14.99", "Conversion rate", "", "2026-02-14"],
        ["ET-012", "SIGNAL", "singapore", "monthly_sub_price_usd", "Subscription", "4.99", "4.99", "No change", "", "2026-02-14"],
        ["ET-013", "VANGUARD", "singapore", "season_length_days", "Season", "90", "75", "Content cadence", "", "2026-02-13"],
        ["ET-014", "SIGNAL", "singapore", "ad_watch_gem_reward", "Ads", "10", "15", "Ad engagement", "", "2026-02-13"],
        ["ET-015", "VANGUARD", "singapore", "crate_drop_rate_legendary", "Drops", "2%", "3%", "Engagement metric", "", "2026-02-12"],
        ["ET-016", "SIGNAL", "singapore", "energy_cap", "Energy", "120", "150", "Session length", "", "2026-02-12"],
        ["ET-017", "VANGUARD", "singapore", "rank_decay_days", "Ranked", "14", "7", "Active play incentive", "", "2026-02-11"],
        ["ET-018", "SIGNAL", "singapore", "gacha_pity_counter", "Gacha", "90", "80", "Player sentiment", "ep-gs-002", "2026-02-19"],
        ["ET-019", "VANGUARD", "singapore", "daily_challenge_reward", "Challenges", "50 currency", "75 currency", "Engagement", "", "2026-02-10"],
        ["ET-020", "SIGNAL", "singapore", "pvp_reward_scaling", "PvP", "1.0x", "1.3x", "PvP participation", "", "2026-02-10"],
        ["ET-021", "VANGUARD", "singapore", "premium_currency_per_usd", "Purchase", "100", "110", "Value perception", "", "2026-02-09"],
        ["ET-022", "SIGNAL", "singapore", "craft_material_cost", "Crafting", "500", "350", "Crafting accessibility", "", "2026-02-09"],
        ["ET-023", "VANGUARD", "singapore", "tournament_entry_fee", "Esports", "100 currency", "50 currency", "Tournament participation", "", "2026-02-08"],
        ["ET-024", "SIGNAL", "singapore", "character_dupe_compensation", "Gacha", "10 shards", "25 shards", "Dupe value", "", "2026-02-08"],
        ["ET-025", "VANGUARD", "singapore", "clan_war_reward_pool", "Social", "10000", "15000", "Clan engagement", "", "2026-02-07"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblEconomyTuning", len(rows))


def _build_feature_cuts(wb: Workbook) -> None:
    """FeatureCuts tab — 25 rows of feature cut/defer/descope decisions."""
    ws = wb.create_sheet("FeatureCuts")
    headers = ["CutID", "Title", "Studio", "Domain", "Feature",
               "Status", "Reason", "Impact", "EpisodeRef", "DecidedBy", "Date"]
    ws.append(headers)

    rows = [
        ["FC-001", "RONIN", "tokyo", "CRE", "Dismemberment VFX system", "Deferred", "PEGI16 envelope breach", "DLC ships without signature feature", "ep-gs-001", "governance_council", "2026-02-19"],
        ["FC-002", "RONIN", "tokyo", "CRE", "Dynamic gore system", "Cut", "Rating risk too high", "Simplified blood effects", "ep-gs-001", "creative_director", "2026-02-19"],
        ["FC-003", "VANGUARD", "singapore", "MON", "Paid random rewards (BE/NL)", "Disabled", "Lootbox prohibition", "Revenue loss in 2 markets", "ep-gs-002", "legal_counsel", "2026-02-19"],
        ["FC-004", "SIGNAL", "singapore", "MON", "Hidden gacha rates", "Replaced", "JP/KR disclosure laws", "Transparent probability UI", "ep-gs-002", "compliance_eng", "2026-02-19"],
        ["FC-005", "VANGUARD", "tokyo", "CRE", "Cross-play voice chat", "Deferred", "Platform cert complexity", "Text chat only at launch", "", "product_manager", "2026-02-18"],
        ["FC-006", "RONIN", "montreal", "CRE", "Photo mode sharing", "Descoped", "Privacy review pending", "Local screenshots only", "", "privacy_eng", "2026-02-17"],
        ["FC-007", "SIGNAL", "montreal", "DAT", "Social graph export", "Cut", "GDPR scope unclear", "Feature removed entirely", "", "privacy_eng", "2026-02-17"],
        ["FC-008", "VANGUARD", "bucharest", "PLT", "Ray tracing (Switch port)", "Cut", "Hardware limitation", "Baked lighting fallback", "", "tech_director", "2026-02-16"],
        ["FC-009", "RONIN", "bucharest", "PLT", "120fps mode (PS5)", "Deferred", "Cert test window conflict", "Ships in patch 1.1", "", "perf_lead", "2026-02-16"],
        ["FC-010", "SIGNAL", "singapore", "OPS", "Real-time PvP matchmaking", "Descoped", "Server cost", "Async PvP at launch", "", "server_eng", "2026-02-15"],
        ["FC-011", "VANGUARD", "tokyo", "CRE", "Destructible environments (ranked)", "Cut", "Competitive integrity", "Cosmetic destruction only", "", "competitive_lead", "2026-02-15"],
        ["FC-012", "RONIN", "tokyo", "CRE", "NPC companion AI v2", "Deferred", "Scope overflow", "v1 AI ships; v2 in DLC2", "", "ai_lead", "2026-02-14"],
        ["FC-013", "VANGUARD", "singapore", "OPS", "Custom server browser", "Cut", "Matchmaking priority", "Official playlists only", "", "product_manager", "2026-02-14"],
        ["FC-014", "SIGNAL", "montreal", "CRE", "User-generated content tools", "Cut", "Moderation risk", "Curated content only", "", "community_lead", "2026-02-13"],
        ["FC-015", "RONIN", "montreal", "DAT", "Cross-platform save sync", "Deferred", "Data residency conflict", "Per-platform saves at launch", "", "privacy_eng", "2026-02-13"],
        ["FC-016", "VANGUARD", "bucharest", "REG", "China-specific social features", "Deferred", "NPPA review pending", "Base features only for CN", "ep-gs-004", "compliance_eng", "2026-02-19"],
        ["FC-017", "SIGNAL", "singapore", "MON", "Gifting system", "Deferred", "Tax implications per region", "Direct purchase only", "", "finance_lead", "2026-02-12"],
        ["FC-018", "RONIN", "tokyo", "CRE", "Extended finisher animations", "Descoped", "Rating boundary", "Shortened finishers", "ep-gs-001", "art_lead", "2026-02-19"],
        ["FC-019", "VANGUARD", "tokyo", "OPS", "Replay system", "Deferred", "Storage cost", "Ships in Season 2", "", "server_eng", "2026-02-11"],
        ["FC-020", "SIGNAL", "singapore", "OPS", "Live event modding tools", "Cut", "Security review", "Internal tools only", "", "security_lead", "2026-02-11"],
        ["FC-021", "RONIN", "bucharest", "PLT", "HDR10+ support", "Deferred", "Cert test backlog", "SDR at launch", "", "graphics_eng", "2026-02-10"],
        ["FC-022", "VANGUARD", "bucharest", "PLT", "Haptic feedback (PS5 DualSense)", "Descoped", "Dev time vs cert window", "Basic rumble only", "", "input_eng", "2026-02-10"],
        ["FC-023", "SIGNAL", "montreal", "CRE", "Seasonal narrative arcs", "Descoped", "Writing team capacity", "Static event themes", "", "narrative_director", "2026-02-09"],
        ["FC-024", "RONIN", "montreal", "CRE", "Alternate language VO (full)", "Deferred", "Recording schedule", "Subtitles + EN/JP VO at launch", "", "localization_lead", "2026-02-09"],
        ["FC-025", "VANGUARD", "singapore", "MON", "Auction house", "Cut", "Economy destabilization risk", "Fixed-price store only", "", "economy_designer", "2026-02-08"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblFeatureCuts", len(rows))


def _build_assumptions(wb: Workbook) -> None:
    """Assumptions tab — 25 rows of institutional assumptions with TTL."""
    ws = wb.create_sheet("Assumptions")
    headers = ["AssumptionID", "Title", "Domain", "Assumption",
               "Confidence", "TTL_Days", "LastValidated", "Source",
               "RiskIfWrong", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["A-001", "RONIN", "REG", "PEGI 16 rating will hold through DLC cycle", "0.60", "30", "2026-02-19", "S018", "Retail recall + re-rating", "ep-gs-001"],
        ["A-002", "VANGUARD", "MON", "No-pay-to-win commitment still holds", "0.55", "14", "2026-02-19", "S005", "Community backlash + press", "ep-gs-002"],
        ["A-003", "ALL", "PLT", "Sony TRC v5.2 requirements unchanged", "0.90", "90", "2026-02-15", "S017", "Cert failure", ""],
        ["A-004", "ALL", "PLT", "Microsoft XR requirements unchanged", "0.90", "90", "2026-02-15", "S017", "Cert failure", ""],
        ["A-005", "RONIN", "PLT", "Nintendo Lotcheck acceptance for Switch SKU", "0.70", "30", "2026-02-10", "S017", "Distribution block", ""],
        ["A-006", "ALL", "OPS", "Cross-studio build pipeline is independent per title", "0.40", "7", "2026-02-19", "S003", "Cascade failure across titles", "ep-gs-003"],
        ["A-007", "ALL", "OPS", "Player telemetry pipeline has no single point of failure", "0.45", "7", "2026-02-19", "S023", "Blind spot across 3 domains", "ep-gs-003"],
        ["A-008", "SIGNAL", "REG", "Japan gacha disclosure rates match implementation", "0.75", "14", "2026-02-19", "S014", "Regulatory action", "ep-gs-002"],
        ["A-009", "VANGUARD", "REG", "Belgium/NL lootbox prohibition scope unchanged", "0.85", "60", "2026-02-01", "S025", "Enforcement action", "ep-gs-002"],
        ["A-010", "ALL", "DAT", "GDPR DPA guidance stable for next quarter", "0.80", "90", "2026-01-15", "S006", "Processing suspension", ""],
        ["A-011", "SIGNAL", "DAT", "China PIPL requirements fully implemented", "0.78", "30", "2026-02-10", "S025", "Market access revocation", ""],
        ["A-012", "ALL", "OPS", "Tokyo-Montreal handoff gap < 6 hours", "0.65", "7", "2026-02-19", "S002", "Overnight drift accumulation", "ep-gs-004"],
        ["A-013", "VANGUARD", "CRE", "Competitive integrity unaffected by monetization", "0.50", "14", "2026-02-19", "S028", "Player trust erosion", "ep-gs-002"],
        ["A-014", "RONIN", "CRE", "Art direction within rating envelope post-DLC", "0.65", "30", "2026-02-19", "S001", "Re-rating required", "ep-gs-001"],
        ["A-015", "ALL", "PLT", "Steam Deck verification requirements unchanged", "0.85", "60", "2026-02-05", "S017", "Delisting risk", ""],
        ["A-016", "SIGNAL", "MON", "Mobile ad revenue projections accurate", "0.70", "14", "2026-02-10", "S014", "Revenue shortfall", ""],
        ["A-017", "VANGUARD", "OPS", "Anti-cheat effectiveness > 95%", "0.80", "7", "2026-02-18", "S013", "Competitive integrity loss", ""],
        ["A-018", "ALL", "REG", "No new lootbox legislation in active markets", "0.60", "30", "2026-02-01", "S025", "Emergency compliance work", ""],
        ["A-019", "RONIN", "CRE", "DLC2 dismemberment will pass re-rating", "0.50", "60", "2026-02-19", "S018", "Feature permanently cut", "ep-gs-001"],
        ["A-020", "VANGUARD", "PLT", "Cross-play cert requirements aligned across platforms", "0.75", "30", "2026-02-12", "S017", "Platform-specific builds needed", ""],
        ["A-021", "ALL", "DAT", "Brazil LGPD enforcement posture unchanged", "0.70", "60", "2026-01-20", "S025", "Market compliance gap", ""],
        ["A-022", "SIGNAL", "OPS", "Server costs within budget at current DAU", "0.75", "14", "2026-02-15", "S013", "Budget overrun", ""],
        ["A-023", "ALL", "OPS", "Shared infrastructure correlation risk is managed", "0.35", "7", "2026-02-19", "S003", "Multi-domain cascade", "ep-gs-003"],
        ["A-024", "VANGUARD", "REG", "South Korea age-gate implementation sufficient", "0.70", "30", "2026-02-10", "S014", "Regulatory fine", ""],
        ["A-025", "RONIN", "REG", "China NPPA approval timeline predictable", "0.55", "30", "2026-02-05", "S025", "China launch delay", ""],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblAssumptions", len(rows))


def _build_drift_signals(wb: Workbook) -> None:
    """DriftSignals tab — 25 rows including the 4 scenario signals."""
    ws = wb.create_sheet("DriftSignals")
    headers = ["DriftID", "Title", "Domain", "Category", "Severity",
               "ClaimsAffected", "EvidenceCount", "DetectedAt",
               "Status", "PatchRef", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["DS-GS-001", "RONIN", "CRE/REG/PLT", "content_rating_mismatch", "RED", "CRE-001,REG-001,PLT-001", "18", "2026-02-19T07:30:00Z", "Patched", "Patch-GS-001", "ep-gs-001"],
        ["DS-GS-002", "VANGUARD", "MON/REG/CRE", "monetization_contradiction", "RED", "MON-001,REG-002,CRE-002", "22", "2026-02-19T09:00:00Z", "Patched", "Patch-GS-002", "ep-gs-002"],
        ["DS-GS-003", "ALL", "OPS/PLT/DAT/MON", "infrastructure_cascade", "RED", "8 claims across 4 domains", "50", "2026-02-19T11:00:00Z", "Patched", "Patch-GS-003", "ep-gs-003"],
        ["DS-GS-004", "VANGUARD", "OPS/PLT", "timezone_regression", "YELLOW", "OPS-001,PLT-001", "12", "2026-02-19T14:00:00Z", "Patched", "Patch-GS-004", "ep-gs-004"],
        ["DS-GS-005", "SIGNAL", "MON", "gacha_rate_drift", "YELLOW", "MON-001b", "4", "2026-02-18T16:00:00Z", "Active", "", ""],
        ["DS-GS-006", "RONIN", "PLT", "lotcheck_resubmission_overdue", "YELLOW", "PLT-001c", "3", "2026-02-18T09:00:00Z", "Active", "", ""],
        ["DS-GS-007", "VANGUARD", "DAT", "telemetry_consent_gap", "YELLOW", "DAT-001a", "6", "2026-02-17T14:00:00Z", "Monitoring", "", ""],
        ["DS-GS-008", "SIGNAL", "REG", "kr_age_gate_bypass", "RED", "REG-002d", "2", "2026-02-17T10:00:00Z", "Active", "", ""],
        ["DS-GS-009", "RONIN", "CRE", "audio_descriptor_mismatch", "GREEN", "CRE-001a", "3", "2026-02-16T15:00:00Z", "Resolved", "", ""],
        ["DS-GS-010", "VANGUARD", "OPS", "matchmaking_skill_compression", "GREEN", "OPS-002", "5", "2026-02-16T12:00:00Z", "Monitoring", "", ""],
        ["DS-GS-011", "ALL", "PLT", "cert_test_backlog_growing", "YELLOW", "PLT-003,PLT-004", "8", "2026-02-15T11:00:00Z", "Active", "", ""],
        ["DS-GS-012", "SIGNAL", "MON", "ad_revenue_below_projection", "GREEN", "MON-004", "4", "2026-02-15T09:00:00Z", "Monitoring", "", ""],
        ["DS-GS-013", "VANGUARD", "CRE", "skin_lore_inconsistency", "GREEN", "CRE-003", "2", "2026-02-14T16:00:00Z", "Resolved", "", ""],
        ["DS-GS-014", "RONIN", "DAT", "replay_retention_exceeded", "YELLOW", "DAT-003", "4", "2026-02-14T10:00:00Z", "Active", "", ""],
        ["DS-GS-015", "ALL", "OPS", "shared_pipeline_latency_spike", "YELLOW", "OPS-001", "12", "2026-02-13T08:00:00Z", "Resolved", "", ""],
        ["DS-GS-016", "SIGNAL", "REG", "cn_playtime_cap_misconfigured", "RED", "REG-003", "5", "2026-02-13T06:00:00Z", "Patched", "", ""],
        ["DS-GS-017", "VANGUARD", "MON", "battle_pass_value_complaint", "GREEN", "MON-003", "3", "2026-02-12T14:00:00Z", "Monitoring", "", ""],
        ["DS-GS-018", "RONIN", "PLT", "ps5_memory_budget_exceeded", "YELLOW", "PLT-002", "6", "2026-02-12T11:00:00Z", "Resolved", "", ""],
        ["DS-GS-019", "SIGNAL", "OPS", "server_cost_overrun_projected", "YELLOW", "OPS-003", "4", "2026-02-11T15:00:00Z", "Active", "", ""],
        ["DS-GS-020", "VANGUARD", "REG", "brazil_lgpd_gap_detected", "YELLOW", "REG-004,DAT-004", "5", "2026-02-11T10:00:00Z", "Active", "", ""],
        ["DS-GS-021", "RONIN", "CRE", "localization_tone_drift", "GREEN", "CRE-004", "3", "2026-02-10T16:00:00Z", "Monitoring", "", ""],
        ["DS-GS-022", "ALL", "OPS", "beacon_sync_variance_elevated", "YELLOW", "OPS-001", "8", "2026-02-10T08:00:00Z", "Monitoring", "", ""],
        ["DS-GS-023", "VANGUARD", "CRE", "esports_rule_inconsistency", "GREEN", "CRE-005", "2", "2026-02-09T14:00:00Z", "Resolved", "", ""],
        ["DS-GS-024", "SIGNAL", "DAT", "pipl_audit_gap", "YELLOW", "DAT-001c", "4", "2026-02-09T09:00:00Z", "Active", "", ""],
        ["DS-GS-025", "RONIN", "REG", "esrb_submission_stale", "YELLOW", "REG-001a", "3", "2026-02-08T11:00:00Z", "Active", "", ""],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblDriftSignals", len(rows))


def _build_patch_plans(wb: Workbook) -> None:
    """PatchPlans tab — 25 rows of patch plan records."""
    ws = wb.create_sheet("PatchPlans")
    headers = ["PatchID", "Title", "DriftRef", "Severity", "Domains",
               "SelectedOption", "Steps", "Owner", "Status",
               "ClosureConditions", "EpisodeRef"]
    ws.append(headers)

    rows = [
        ["Patch-GS-001", "RONIN DLC rating envelope", "DS-GS-001", "RED", "CRE/REG/PLT", "C", "3", "tokyo", "Applied", "REG-001>=0.85; PLT-001 gates unchanged; no RED CRE/REG/PLT", "ep-gs-001"],
        ["Patch-GS-002", "Founder's Cache monetization", "DS-GS-002", "RED", "MON/REG/CRE", "A", "3", "singapore", "Applied", "MON-001 restored; REG-002 restored", "ep-gs-002"],
        ["Patch-GS-003", "Shared infra concentration", "DS-GS-003", "RED", "OPS/PLT/DAT/MON", "-", "3", "tokyo", "Applied", "shared-infra < 40 nodes; Tier0 margin improved", "ep-gs-003"],
        ["Patch-GS-004", "Timezone regression", "DS-GS-004", "YELLOW", "OPS/PLT", "-", "2", "montreal", "Applied", "No regression 14d; PLT stable across handoffs", "ep-gs-004"],
        ["Patch-GS-005", "Gacha rate realignment", "DS-GS-005", "YELLOW", "MON", "-", "1", "singapore", "Planned", "Published rates match server config", ""],
        ["Patch-GS-006", "Lotcheck resubmission", "DS-GS-006", "YELLOW", "PLT", "-", "2", "bucharest", "In Progress", "Nintendo acceptance received", ""],
        ["Patch-GS-007", "Telemetry consent fix", "DS-GS-007", "YELLOW", "DAT", "-", "1", "montreal", "Planned", "Opt-in default for all regions", ""],
        ["Patch-GS-008", "KR age-gate enforcement", "DS-GS-008", "RED", "REG", "-", "2", "singapore", "In Progress", "Age gate passes all KR test cases", ""],
        ["Patch-GS-009", "CN playtime fix", "DS-GS-016", "RED", "REG", "-", "1", "bucharest", "Applied", "NPPA cap correctly enforced", ""],
        ["Patch-GS-010", "PS5 memory optimization", "DS-GS-018", "YELLOW", "PLT", "-", "2", "bucharest", "Applied", "Memory within Sony TRC limits", ""],
        ["Patch-GS-011", "LGPD compliance gap", "DS-GS-020", "YELLOW", "REG/DAT", "-", "2", "montreal", "Planned", "Brazil data handling compliant", ""],
        ["Patch-GS-012", "Cert backlog reduction", "DS-GS-011", "YELLOW", "PLT", "-", "3", "bucharest", "In Progress", "Backlog < 5 items", ""],
        ["Patch-GS-013", "Replay retention cleanup", "DS-GS-014", "YELLOW", "DAT", "-", "1", "montreal", "Planned", "Retention within policy", ""],
        ["Patch-GS-014", "Server cost optimization", "DS-GS-019", "YELLOW", "OPS", "-", "2", "singapore", "Planned", "Costs within quarterly budget", ""],
        ["Patch-GS-015", "PIPL audit remediation", "DS-GS-024", "YELLOW", "DAT", "-", "2", "montreal", "Planned", "Audit gap closed", ""],
        ["Patch-GS-016", "ESRB submission refresh", "DS-GS-025", "YELLOW", "REG", "-", "1", "bucharest", "Planned", "Submission current", ""],
        ["Patch-GS-017", "Beacon sync stabilization", "DS-GS-022", "YELLOW", "OPS", "-", "1", "tokyo", "Monitoring", "Variance within threshold", ""],
        ["Patch-GS-018", "Build pipeline isolation", "DS-GS-003", "RED", "OPS", "-", "3", "tokyo", "In Progress", "Per-title CI isolation", ""],
        ["Patch-GS-019", "Telemetry pipeline redundancy", "DS-GS-003", "RED", "OPS", "-", "2", "montreal", "In Progress", "Secondary pipeline active", ""],
        ["Patch-GS-020", "Battle pass value rebalance", "DS-GS-017", "GREEN", "MON", "-", "1", "singapore", "Monitoring", "Sentiment stabilized", ""],
        ["Patch-GS-021", "Localization tone review", "DS-GS-021", "GREEN", "CRE", "-", "1", "montreal", "Monitoring", "Tone guide updated", ""],
        ["Patch-GS-022", "Esports rule sync", "DS-GS-023", "GREEN", "CRE", "-", "1", "tokyo", "Applied", "Rules consistent across regions", ""],
        ["Patch-GS-023", "Ad revenue model update", "DS-GS-012", "GREEN", "MON", "-", "1", "singapore", "Monitoring", "Projections recalibrated", ""],
        ["Patch-GS-024", "Skin lore correction", "DS-GS-013", "GREEN", "CRE", "-", "1", "tokyo", "Applied", "Lore database updated", ""],
        ["Patch-GS-025", "Pipeline latency fix", "DS-GS-015", "YELLOW", "OPS", "-", "2", "tokyo", "Applied", "Latency within SLA", ""],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblPatchPlans", len(rows))


def _build_canon_rules(wb: Workbook) -> None:
    """CanonRules tab — 25 rows of franchise canon/governance rules."""
    ws = wb.create_sheet("CanonRules")
    headers = ["RuleID", "Title", "Domain", "Rule", "Tier",
               "Enforcement", "Source", "LastReviewed"]
    ws.append(headers)

    rows = [
        ["CR-001", "RONIN", "CRE", "All content must fit within approved rating envelope per region", "0", "Blocking", "S018", "2026-02-19"],
        ["CR-002", "RONIN", "CRE", "Art direction changes require Tokyo creative director sign-off", "0", "Blocking", "S001", "2026-02-15"],
        ["CR-003", "VANGUARD", "CRE", "No gameplay mechanics that provide competitive advantage for purchase", "0", "Blocking", "S005", "2026-02-19"],
        ["CR-004", "SIGNAL", "CRE", "Visual content must be appropriate for E10+/PEGI 7 at all times", "0", "Blocking", "S018", "2026-02-15"],
        ["CR-005", "ALL", "REG", "Regional rating must be current and valid before distribution", "0", "Blocking", "S021", "2026-02-15"],
        ["CR-006", "ALL", "REG", "Lootbox/gacha must comply with per-country legislation", "0", "Blocking", "S025", "2026-02-19"],
        ["CR-007", "ALL", "REG", "China distribution requires current NPPA approval", "0", "Blocking", "S025", "2026-02-05"],
        ["CR-008", "ALL", "PLT", "Platform cert must be current before any store update", "0", "Blocking", "S017", "2026-02-15"],
        ["CR-009", "ALL", "PLT", "No cert regression in compliance patches", "1", "Gating", "S010", "2026-02-19"],
        ["CR-010", "ALL", "MON", "Published monetization rates must match server implementation", "0", "Blocking", "S014", "2026-02-19"],
        ["CR-011", "ALL", "MON", "No surprise price increases without 30-day notice", "1", "Gating", "S005", "2026-02-10"],
        ["CR-012", "ALL", "DAT", "Player data handling must comply with jurisdiction of residence", "0", "Blocking", "S006", "2026-02-15"],
        ["CR-013", "ALL", "DAT", "Data deletion requests must complete within regulatory timeframe", "0", "Blocking", "S006", "2026-02-15"],
        ["CR-014", "ALL", "OPS", "Live service patches require regression testing before deploy", "1", "Gating", "S009", "2026-02-18"],
        ["CR-015", "ALL", "OPS", "Server-side config changes must have rollback plan documented", "1", "Gating", "S013", "2026-02-17"],
        ["CR-016", "RONIN", "CRE", "Franchise lore changes require narrative review board approval", "1", "Gating", "S005", "2026-02-14"],
        ["CR-017", "VANGUARD", "CRE", "Weapon balance changes must pass competitive review", "1", "Gating", "S001", "2026-02-14"],
        ["CR-018", "SIGNAL", "MON", "Gacha pity system must be active in all markets", "1", "Gating", "S014", "2026-02-12"],
        ["CR-019", "ALL", "OPS", "Cross-timezone handoffs must include status beacon ping", "2", "Advisory", "S002", "2026-02-19"],
        ["CR-020", "ALL", "PLT", "Accessibility features must meet platform-specific requirements", "2", "Advisory", "S017", "2026-02-10"],
        ["CR-021", "ALL", "REG", "Marketing materials must be consistent with current ratings", "1", "Gating", "S018", "2026-02-19"],
        ["CR-022", "VANGUARD", "OPS", "Ranked season resets require 48h advance notice", "2", "Advisory", "S013", "2026-02-08"],
        ["CR-023", "SIGNAL", "OPS", "Event schedules must account for all active timezones", "2", "Advisory", "S013", "2026-02-08"],
        ["CR-024", "ALL", "DAT", "Telemetry collection requires explicit opt-in in GDPR regions", "0", "Blocking", "S006", "2026-02-15"],
        ["CR-025", "ALL", "CRE", "Cross-title franchise consistency reviewed quarterly", "2", "Advisory", "S001", "2026-01-15"],
    ]
    for row in rows:
        ws.append(row)
    _add_table(ws, "tblCanonRules", len(rows))


def _build_prompts(wb: Workbook) -> None:
    """PROMPTS tab — LLM interaction surface."""
    ws = wb.create_sheet("PROMPTS")

    ws["A1"] = "What would you like to do today?"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Options:"
    ws["A4"] = "(1) Why did we change X?"
    ws["A5"] = "(2) What drifted?"
    ws["A6"] = "(3) Generate patch plan"
    ws["A7"] = "(4) Audit compliance loop"
    ws["A8"] = "(5) Show blast radius for a source"
    ws["A9"] = "(6) Score current lattice health"

    ws["A11"] = "System Prompt (for LLM integration):"
    ws["A11"].font = Font(bold=True)

    system_prompt = (
        "You are a GameOps governance assistant for Nexus Interactive (fictional). "
        "Read ALL tabs in this workbook: BalanceChanges, EconomyTuning, FeatureCuts, "
        "Assumptions, DriftSignals, PatchPlans, CanonRules. Cross-reference episode IDs "
        "(ep-gs-XXX), drift IDs (DS-GS-XXX), and patch IDs (Patch-GS-XXX) across tabs. "
        "When the user asks a question:\n"
        "1. Identify which tabs contain relevant data\n"
        "2. Trace the decision chain: BalanceChange -> DriftSignal -> PatchPlan -> Episode\n"
        "3. Check CanonRules for any governance violations\n"
        "4. Check Assumptions for any expired or low-confidence beliefs\n"
        "5. Respond with: (a) what happened, (b) why it matters, (c) what to do next\n"
        "6. Link to specific row IDs so the user can verify\n\n"
        "Studios: Tokyo, Montreal, Bucharest, Singapore\n"
        "Titles: RONIN, VANGUARD, SIGNAL\n"
        "Domains: CRE, REG, PLT, MON, OPS, DAT"
    )
    ws["A12"] = system_prompt

    # Set column width for readability
    ws.column_dimensions["A"].width = 100


def generate_workbook(output_path: Path) -> None:
    """Generate the complete GameOps Pack workbook."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    _build_balance_changes(wb)
    _build_economy_tuning(wb)
    _build_feature_cuts(wb)
    _build_assumptions(wb)
    _build_drift_signals(wb)
    _build_patch_plans(wb)
    _build_canon_rules(wb)
    _build_prompts(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"GameOps Workbook generated: {output_path}")
    print(f"  Tabs: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
    print(f"  Rows: {25 * 7 + 12} (25 data rows x 7 tables + PROMPTS)")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate GameOps Pack Excel workbook",
    )
    parser.add_argument(
        "--out",
        type=str,
        default="./GameOps_Workbook.xlsx",
        help="Output path for the workbook (default: ./GameOps_Workbook.xlsx)",
    )
    args = parser.parse_args()
    generate_workbook(Path(args.out))


if __name__ == "__main__":
    main()
