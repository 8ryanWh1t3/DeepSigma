#!/usr/bin/env python3
"""Validate that an Excel workbook conforms to the WORKBOOK BOOT Protocol.

Rules enforced:
  1. Workbook contains sheet named exactly "BOOT"
  2. Cell BOOT!A1 is non-empty and begins with "BOOT!"
  3. BOOT!A1 contains required metadata keys (case-insensitive):
     version:, ttl_hours_default:, risk_lane_default:, schema_ref:, owner:
  4. Workbook contains 7 canonical governance named tables (unless --boot-only)

Usage:
    python tools/validate_workbook_boot.py <path.xlsx> [--boot-only]

Exit codes:
    0 = pass
    1 = validation failure
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

REQUIRED_KEYS = [
    "version:",
    "ttl_hours_default:",
    "risk_lane_default:",
    "schema_ref:",
    "owner:",
]

REQUIRED_TABLE_FRAGMENTS = [
    "Timeline",
    "Deliverables",
    "DLR",
    "Claims",
    "Assumptions",
    "PatchLog",
    "Guardrails",
]


def validate(workbook_path: str, *, boot_only: bool = False) -> list[str]:
    """Validate a workbook against the BOOT protocol.

    Returns a list of error strings. Empty list = pass.
    """
    import openpyxl

    errors: list[str] = []
    path = Path(workbook_path)

    if not path.exists():
        return [f"File not found: {path}"]

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Rule 1: BOOT sheet exists
    if "BOOT" not in wb.sheetnames:
        errors.append("Missing required sheet: BOOT")
        wb.close()
        return errors  # Cannot proceed without BOOT sheet

    # Rule 2: BOOT!A1 non-empty, starts with "BOOT!"
    boot_ws = wb["BOOT"]
    a1_value = boot_ws["A1"].value
    if not a1_value:
        errors.append("BOOT!A1 is empty")
        wb.close()
        return errors

    a1_text = str(a1_value)
    if not a1_text.startswith("BOOT!"):
        errors.append(
            f"BOOT!A1 must start with 'BOOT!' prefix, "
            f"got: {a1_text[:40]!r}..."
        )

    # Rule 3: Required metadata keys
    a1_lower = a1_text.lower()
    for key in REQUIRED_KEYS:
        if key.lower() not in a1_lower:
            errors.append(f"BOOT!A1 missing required key: {key}")

    # Rule 4: Named tables (unless --boot-only)
    if not boot_only:
        # Need read_only=False to access tables
        wb.close()
        wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)
        found_tables: set[str] = set()
        for ws in wb.worksheets:
            for table in ws.tables.values():
                found_tables.add(table.displayName)

        for fragment in REQUIRED_TABLE_FRAGMENTS:
            matched = any(fragment in name for name in found_tables)
            if not matched:
                errors.append(
                    f"Missing named table containing '{fragment}' "
                    f"(found: {sorted(found_tables) if found_tables else 'none'})"
                )

    wb.close()
    return errors


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate workbook BOOT protocol conformance",
    )
    parser.add_argument("workbook", help="Path to .xlsx workbook")
    parser.add_argument(
        "--boot-only",
        action="store_true",
        help="Only validate BOOT sheet (skip named table checks)",
    )
    args = parser.parse_args(argv)

    errors = validate(args.workbook, boot_only=args.boot_only)

    if errors:
        print(f"FAIL — {len(errors)} validation error(s):")
        for e in errors:
            print(f"  - {e}")
        return 1

    print(f"PASS — {args.workbook}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
