#!/usr/bin/env python3
"""Generate small .xlsx test fixtures for BOOT contract validation tests.

Run this script once to regenerate the fixture files.

Usage:
    python tests/fixtures/generate_test_workbooks.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo

FIXTURES_DIR = Path(__file__).resolve().parent

VALID_BOOT_A1 = """\
BOOT! METADATA
version: 0.6.3
ttl_hours_default: 168
risk_lane_default: ADVISORY
schema_ref: docs/excel-first/TABLE_SCHEMAS.md
owner: Test Fixture Team

=== TEST WORKBOOK ===
YOU ARE: Test Copilot
"""

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)

TABLE_DEFS = [
    ("tblTimeline", ["Week", "Phase", "Status"]),
    ("tblDeliverables", ["Asset_ID", "Type", "Status"]),
    ("tblDLR", ["Decision_ID", "Context", "Status"]),
    ("tblClaims", ["Claim_ID", "Text", "Confidence"]),
    ("tblAssumptions", ["Assumption_ID", "Text", "Status"]),
    ("tblPatchLog", ["Patch_ID", "Source", "Status"]),
    ("tblCanonGuardrails", ["Guardrail_ID", "Rule", "Severity"]),
]


def _add_table(wb, sheet_name, table_name, columns, sample_row):
    ws = wb.create_sheet(sheet_name)
    for c, header in enumerate(columns, 1):
        ws.cell(row=1, column=c, value=header)
    for c, val in enumerate(sample_row, 1):
        ws.cell(row=c, column=c, value=val)
    # Write one data row
    for c, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=c, value=val)
    col_letter = openpyxl.utils.get_column_letter(len(columns))
    ref = f"A1:{col_letter}2"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)


def generate_valid_boot():
    """Full valid workbook with BOOT + 7 named tables."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    boot = wb.create_sheet("BOOT")
    boot["A1"] = VALID_BOOT_A1
    for tbl_name, cols in TABLE_DEFS:
        sheet_name = tbl_name.replace("tbl", "")
        sample = [f"test-{i}" for i in range(len(cols))]
        _add_table(wb, sheet_name, tbl_name, cols, sample)
    wb.save(str(FIXTURES_DIR / "valid_boot.xlsx"))
    print("Created valid_boot.xlsx")


def generate_missing_boot_sheet():
    """Workbook with no BOOT sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA"
    ws["A1"] = "This workbook has no BOOT sheet"
    wb.save(str(FIXTURES_DIR / "missing_boot_sheet.xlsx"))
    print("Created missing_boot_sheet.xlsx")


def generate_missing_keys():
    """BOOT sheet present but A1 is missing required metadata keys."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    boot = wb.create_sheet("BOOT")
    boot["A1"] = "BOOT! METADATA\nversion: 0.6.3\n\nMissing other keys."
    wb.save(str(FIXTURES_DIR / "missing_keys.xlsx"))
    print("Created missing_keys.xlsx")


def generate_boot_only():
    """Valid BOOT sheet, but no named tables."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    boot = wb.create_sheet("BOOT")
    boot["A1"] = VALID_BOOT_A1
    wb.save(str(FIXTURES_DIR / "boot_only.xlsx"))
    print("Created boot_only.xlsx")


if __name__ == "__main__":
    generate_valid_boot()
    generate_missing_boot_sheet()
    generate_missing_keys()
    generate_boot_only()
    print(f"\nAll fixtures written to {FIXTURES_DIR}")
