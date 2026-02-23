"""Excel-first Money Demo pipeline — deterministic Drift→Patch artifacts.

Steps:
  1. Generate  — produce workbook.xlsx
  2. Validate  — run BOOT contract validator
  3. Run       — pick scenario row, produce run_record.json
  4. Drift     — simulate TTL expiry, emit drift_signal.json
  5. Patch     — emit patch_stub.json
  6. Delta     — compute coherence_delta.txt

All outputs are deterministic. No LLM calls. No network.
"""

from __future__ import annotations

import json
from pathlib import Path


def run_demo(output_dir: str) -> dict:
    """Run the full demo pipeline. Returns summary dict."""
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    workbook_path = out / "workbook.xlsx"

    # Step 1: Generate workbook
    print("[1/6] Generating workbook...")
    _generate_workbook(workbook_path)

    # Step 2: Validate BOOT contract
    print("[2/6] Validating BOOT contract...")
    boot_valid = _validate_boot(workbook_path)

    # Step 3: Simulate run (pick scenario)
    print("[3/6] Running scenario (ASM-005)...")
    run_record = _simulate_run()
    _write_json(out / "run_record.json", run_record)

    # Step 4: Detect drift
    print("[4/6] Detecting drift...")
    drift_signal = _detect_drift(run_record)
    _write_json(out / "drift_signal.json", drift_signal)

    # Step 5: Propose patch
    print("[5/6] Proposing patch...")
    patch_stub = _propose_patch(drift_signal)
    _write_json(out / "patch_stub.json", patch_stub)

    # Step 6: Compute coherence delta
    print("[6/6] Computing coherence delta...")
    before_score = _compute_score(run_record)
    after_score = _compute_score_after_patch(run_record, patch_stub)
    delta_text = _format_delta(before_score, after_score, drift_signal, patch_stub)
    (out / "coherence_delta.txt").write_text(delta_text)

    return {
        "output_dir": str(out),
        "boot_valid": boot_valid,
        "drift_detected": drift_signal["drift_detected"],
        "patch_proposed": patch_stub["patch_id"] is not None,
        "before_score": before_score,
        "after_score": after_score,
    }


def _generate_workbook(path: Path) -> None:
    """Generate workbook using the CDS generator."""
    import openpyxl
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: F401

    # Import generator functions from the existing tool
    import sys
    from pathlib import Path as P
    repo_root = P(__file__).resolve().parents[3]
    if str(repo_root) not in sys.path:
        sys.path.insert(0, str(repo_root))

    from tools.generate_cds_workbook import (
        BOOT_TEXT, write_rows, add_named_table, style_header_row,
        generate_timeline_rows, generate_deliverables_rows,
        generate_dlr_rows, generate_claims_rows,
        generate_assumptions_rows, generate_patch_rows,
        generate_canon_rows,
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # BOOT sheet
    boot = wb.create_sheet("BOOT")
    boot["A1"] = BOOT_TEXT
    boot["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    boot["A1"].font = Font(name="Consolas", size=11)
    boot.column_dimensions["A"].width = 120

    # Governance tables (smaller — 5 rows for demo speed)
    sheets = [
        ("BRIEF_MATRIX", "tblTimeline", generate_timeline_rows),
        ("DELIVERABLES", "tblDeliverables", generate_deliverables_rows),
        ("DLR_CAPTURE", "tblDLR", generate_dlr_rows),
        ("CLAIMS", "tblClaims", generate_claims_rows),
        ("ASSUMPTIONS", "tblAssumptions", generate_assumptions_rows),
        ("PATCH_LOG", "tblPatchLog", generate_patch_rows),
        ("CANON_SYNC", "tblCanonGuardrails", generate_canon_rows),
    ]

    n_rows = 10  # Smaller than full 25 for demo speed
    for sheet_name, table_name, gen_fn in sheets:
        ws = wb.create_sheet(sheet_name)
        rows = gen_fn(n_rows)
        write_rows(ws, rows)
        style_header_row(ws)
        add_named_table(ws, table_name, n_rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def _validate_boot(workbook_path: Path) -> bool:
    """Validate BOOT contract. Returns True if valid."""
    from tools.validate_workbook_boot import validate
    errors = validate(str(workbook_path))
    if errors:
        for e in errors:
            print(f"  WARN: {e}")
        return False
    return True


def _simulate_run() -> dict:
    """Simulate a scenario run using ASM-005 from the generated data."""
    return {
        "scenario_id": "excel-demo-001",
        "timestamp": "2026-02-19T12:00:00Z",
        "assumption_id": "ASM-005",
        "assumption_text": "Production timeline can absorb 1-week compression",
        "decision_ref": "DEC-005",
        "initial_confidence": 0.8,
        "current_confidence": 0.35,
        "half_life_days": 21,
        "days_since_validation": 42,
        "ttl_expired": True,
        "tables_scanned": [
            "tblTimeline", "tblDeliverables", "tblDLR",
            "tblClaims", "tblAssumptions", "tblPatchLog",
            "tblCanonGuardrails",
        ],
        "required_fields_present": 12,
        "required_fields_total": 14,
        "canon_guardrails_checked": 10,
        "canon_guardrails_passed": 9,
        "canon_guardrails_flagged": 1,
    }


def _detect_drift(run_record: dict) -> dict:
    """Detect drift from run record. Deterministic."""
    return {
        "drift_id": "DRIFT-EXCEL-001",
        "drift_detected": True,
        "drift_type": "freshness",
        "severity": "HIGH",
        "source": run_record["assumption_id"],
        "source_table": "tblAssumptions",
        "decision_ref": run_record["decision_ref"],
        "detected_at": "2026-02-19T12:00:01Z",
        "details": {
            "assumption_text": run_record["assumption_text"],
            "initial_confidence": run_record["initial_confidence"],
            "current_confidence": run_record["current_confidence"],
            "half_life_days": run_record["half_life_days"],
            "days_since_validation": run_record["days_since_validation"],
            "ttl_expired": run_record["ttl_expired"],
        },
        "recommended_patch_type": "RETCON",
        "fingerprint": "excel-demo:freshness:ASM-005:v1",
    }


def _propose_patch(drift_signal: dict) -> dict:
    """Propose a patch based on drift signal. Deterministic."""
    return {
        "patch_id": "PAT-EXCEL-001",
        "drift_ref": drift_signal["drift_id"],
        "decision_ref": drift_signal["decision_ref"],
        "patch_type": drift_signal["recommended_patch_type"],
        "action": "Refresh ASM-005 with current project timeline data",
        "target_table": "tblAssumptions",
        "target_row": drift_signal["source"],
        "fields_updated": {
            "Current_Confidence": 0.75,
            "Date_Validated_Last": "2026-02-19",
            "Status": "REFRESHED",
            "Half_Life_Days": 28,
        },
        "severity": drift_signal["severity"],
        "proposed_at": "2026-02-19T12:00:02Z",
        "proposed_by": "excel-demo-pipeline",
        "status": "PROPOSED",
        "canon_guardrail_check": "GR-005: PASS",
        "expected_ci_impact": "+8 pts",
    }


def _compute_score(run_record: dict) -> float:
    """Compute coherence score from run record fields."""
    field_score = (
        run_record["required_fields_present"]
        / run_record["required_fields_total"]
    ) * 50
    canon_score = (
        run_record["canon_guardrails_passed"]
        / run_record["canon_guardrails_checked"]
    ) * 30
    confidence_score = run_record["current_confidence"] * 20
    return round(field_score + canon_score + confidence_score, 1)


def _compute_score_after_patch(run_record: dict, patch_stub: dict) -> float:
    """Compute score after patch is applied."""
    field_score = (
        run_record["required_fields_present"]
        / run_record["required_fields_total"]
    ) * 50
    canon_score = (
        run_record["canon_guardrails_passed"]
        / run_record["canon_guardrails_checked"]
    ) * 30
    patched_confidence = patch_stub["fields_updated"]["Current_Confidence"]
    confidence_score = patched_confidence * 20
    return round(field_score + canon_score + confidence_score, 1)


def _format_delta(
    before: float, after: float, drift: dict, patch: dict,
) -> str:
    """Format coherence delta report."""
    return f"""\
Coherence Delta Report — Excel-first Money Demo
================================================

Timestamp: 2026-02-19T12:00:03Z
Pipeline:  demos.excel_first

--- BEFORE ---
before_score: {before}
assumptions_expired: 1 (ASM-005)
canon_flags: 1

--- DRIFT ---
drift_id: {drift['drift_id']}
drift_type: {drift['drift_type']}
severity: {drift['severity']}
source: {drift['source']} ({drift['source_table']})
confidence_drop: {drift['details']['initial_confidence']} -> {drift['details']['current_confidence']}

--- PATCH ---
patch_id: {patch['patch_id']}
action: {patch['action']}
target: {patch['target_table']}/{patch['target_row']}
confidence_restored: {patch['fields_updated']['Current_Confidence']}

--- AFTER ---
after_score: {after}
delta: +{after - before:.1f} pts

--- VERDICT ---
Drift detected. Patch proposed. Coherence improved.
"""


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n")
